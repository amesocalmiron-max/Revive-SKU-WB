# =========================
# Stage A: Input + Manifest (WB Necromancer v2)
# Paste into wb_necromancer.py (or run as standalone for now)
# =========================

from __future__ import annotations

import json
import math
import os
import re
import sys
import hashlib
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

SCRIPT_NAME = "wb_necromancer.py"
SCRIPT_VERSION = "2.0.0-dev"
SCHEMA_VERSION = "2.0"

FOCUS_TYPE = "tpu_pocket"
CLUSTERS = ["phone", "type"]

DEFAULT_REVIEW_CFG = {
    "window_days": 90,
    "recent_days": [30, 90],
    "page_size": 30,
    "max_pages": 8,
    "early_stop": True,
    "cache_ttl_hours": 24,
}

DEFAULT_MARKET_THRESHOLDS = {
    # Placeholder thresholds: stored in manifest for reproducibility.
    # We'll tune later with real observed distributions.
    "alive": {"recent_90_min": 8, "days_since_last_max": 30},
    "slow":  {"recent_90_min": 3, "days_since_last_max": 90},
    "dead":  {"recent_90_max": 2, "days_since_last_min": 120},
}

DEFAULT_FILTERS = {
    "tpu_terms": ["tpu", "силикон", "силиконовый", "термополиуретан"],
    "pocket_terms": ["карман", "карт", "карты", "cardholder", "держатель карт"],
    "ban_terms_default": [
        "стекло", "пленка", "плёнка", "зарядка", "кабель", "наушники",
        "ремешок", "часы", "бампер для мебели",
    ],
}

def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def safe_str(x: Any, default: str = "") -> str:
    if x is None:
        return default
    try:
        return str(x)
    except Exception:
        return default

def safe_int(x: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        s = safe_str(x, "").strip()
        if not s:
            return default
        if re.fullmatch(r"-?\d+(\.0+)?", s):
            return int(float(s))
        return int(s)
    except Exception:
        return default

def nm_id_to_str(x: Any) -> str:
    """Coerce nm_id to a stable decimal string (avoid Excel scientific notation issues)."""
    if x is None or isinstance(x, bool):
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if math.isfinite(x):
            return str(int(x))
        return safe_str(x, "").strip()

    s = safe_str(x, "").strip()
    if not s:
        return ""

    if re.fullmatch(r"\d+(\.0+)?", s):
        try:
            return str(int(Decimal(s)))
        except Exception:
            return str(int(float(s)))

    # Scientific notation (Excel)
    if re.fullmatch(r"\d+(\.\d+)?[eE][\+\-]?\d+", s):
        try:
            return str(int(Decimal(s)))
        except (InvalidOperation, ValueError, OverflowError):
            try:
                return str(int(float(s)))
            except Exception:
                return s

    return s

def sha1_short(s: str, n: int = 8) -> str:
    h = hashlib.sha1(s.encode("utf-8"), usedforsecurity=False).hexdigest()
    return h[:n]

def write_json(path: Path, data: Any) -> None:
    """Atomic-ish JSON write: write to temp, then replace."""
    ensure_dir(path.parent)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def scope_hash(scope_list: List[dict]) -> str:
    """Hash only scope identity fields to detect accidental mixing/resume issues."""
    payload = [{
        "nm_id": safe_str(x.get("nm_id"), ""),
        "vendor_code": safe_str(x.get("vendor_code"), ""),
        "name": safe_str(x.get("name"), ""),
        "potential_qty": x.get("potential_qty", None),
    } for x in scope_list]
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha1(raw.encode("utf-8"), usedforsecurity=False).hexdigest()

def read_input_scope(
    input_xlsx: Path,
    sheet: str,
    *,
    expect_count: Optional[int] = None,
    dedupe: bool = False,
) -> List[dict]:
    """Read SKU scope from Excel."""
    if not input_xlsx.exists():
        raise FileNotFoundError(f"Input file not found: {input_xlsx}")

    wb = load_workbook(input_xlsx, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet}. Available: {wb.sheetnames}")
    ws = wb[sheet]

    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        headers.append(safe_str(ws.cell(row=1, column=c).value).strip())

    def col_idx(candidates: List[str]) -> Optional[int]:
        for cand in candidates:
            if cand in headers:
                return headers.index(cand) + 1
        low = [h.lower() for h in headers]
        for cand in candidates:
            cl = cand.lower()
            if cl in low:
                return low.index(cl) + 1
        return None

    c_nm = col_idx(["nm_id", "nmId", "nm"])
    c_vc = col_idx(["vendor_code", "vendorCode", "vc"])
    c_name = col_idx(["name", "title"])
    c_qty = col_idx(["potential_qty", "potentialQty", "qty", "potential"])

    if c_nm is None:
        raise ValueError("Input sheet must contain an 'nm_id' column")

    items: List[dict] = []
    for r in range(2, ws.max_row + 1):
        nm = nm_id_to_str(ws.cell(row=r, column=c_nm).value)
        if not nm:
            continue
        vc = safe_str(ws.cell(row=r, column=c_vc).value).strip() if c_vc else ""
        nm_name = safe_str(ws.cell(row=r, column=c_name).value).strip() if c_name else ""
        qty = safe_int(ws.cell(row=r, column=c_qty).value, None) if c_qty else None
        items.append({"nm_id": nm, "vendor_code": vc, "name": nm_name, "potential_qty": qty})

    if not items:
        raise ValueError(f"No SKU rows found in sheet '{sheet}' (nm_id column is empty?)")

    ids = [x["nm_id"] for x in items]
    if len(set(ids)) != len(ids):
        if not dedupe:
            raise ValueError("Duplicate nm_id in input sheet")
        seen = set()
        deduped: List[dict] = []
        dropped = 0
        for it in items:
            if it["nm_id"] in seen:
                dropped += 1
                continue
            seen.add(it["nm_id"])
            deduped.append(it)
        items = deduped
        print(f"[A] warning: deduped scope, dropped {dropped} duplicate rows", file=sys.stderr)

    if expect_count is not None and len(items) != int(expect_count):
        raise ValueError(f"Expected exactly {int(expect_count)} SKU rows in {sheet}. Got {len(items)}")

    return items

def stage_A(
    out_dir: Path,
    input_xlsx: Path,
    sheet: str,
    *,
    expect_count: Optional[int] = None,
    dedupe: bool = False,
    resume: bool = False,
    review_cfg: Optional[Dict[str, Any]] = None,
    market_thresholds: Optional[Dict[str, Any]] = None,
    filters_cfg: Optional[Dict[str, Any]] = None,
) -> Path:
    """Stage A: create run_manifest.json and initialize run folders."""
    ensure_dir(out_dir)
    ensure_dir(out_dir / ".wb_cache")

    manifest_path = out_dir / "run_manifest.json"

    # Don't silently overwrite an existing run.
    if manifest_path.exists():
        if not resume:
            raise FileExistsError(
                f"Manifest already exists: {manifest_path}. Use --resume to reuse it, or choose a new --out."
            )
        old = read_json(manifest_path)
        if safe_str(old.get("schema_version")) != SCHEMA_VERSION:
            raise ValueError(
                f"Existing manifest schema_version={old.get('schema_version')} does not match expected {SCHEMA_VERSION}."
            )
        old_in = old.get("input", {}) if isinstance(old, dict) else {}
        if Path(safe_str(old_in.get("file", ""))).name != input_xlsx.name or safe_str(old_in.get("sheet")) != sheet:
            raise ValueError(
                "--resume refused: input file/sheet mismatch with existing manifest. Prevents mixing runs by accident."
            )

        scope_list = read_input_scope(input_xlsx, sheet, expect_count=expect_count, dedupe=dedupe)
        new_hash = scope_hash(scope_list)
        if safe_str(old.get("scope_hash")) and safe_str(old.get("scope_hash")) != new_hash:
            raise ValueError("--resume refused: scope_hash differs. Input scope changed since the last run.")

        print(f"[A] resume: manifest OK ({manifest_path})")
        return manifest_path

    scope_list = read_input_scope(input_xlsx, sheet, expect_count=expect_count, dedupe=dedupe)

    rcfg = dict(DEFAULT_REVIEW_CFG)
    if isinstance(review_cfg, dict):
        rcfg.update(review_cfg)

    mthr = dict(DEFAULT_MARKET_THRESHOLDS)
    if isinstance(market_thresholds, dict):
        for k, v in market_thresholds.items():
            if isinstance(v, dict) and isinstance(mthr.get(k), dict):
                tmp = dict(mthr[k])
                tmp.update(v)
                mthr[k] = tmp
            else:
                mthr[k] = v

    fcfg = dict(DEFAULT_FILTERS)
    if isinstance(filters_cfg, dict):
        for k, v in filters_cfg.items():
            fcfg[k] = v

    run_id = f"run_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}_{sha1_short(str(out_dir.resolve()))}"

    manifest = {
        "schema_version": SCHEMA_VERSION,
        "script": {"name": SCRIPT_NAME, "version": SCRIPT_VERSION},
        "created_at": utc_now_iso(),
        "run_id": run_id,
        "project": {
            "name": "WB Revival v2 Necromancer",
            "focus_type": FOCUS_TYPE,
            "clusters": CLUSTERS,
        },
        "input": {
            "file": str(input_xlsx),
            "sheet": sheet,
            "expect_count": int(expect_count) if expect_count is not None else None,
            "dedupe": bool(dedupe),
        },
        "scope": {"sku_list": scope_list},
        "scope_count": len(scope_list),
        "scope_hash": scope_hash(scope_list),
        "config": {
            "reviews": rcfg,
            "market_thresholds": mthr,
            "filters": fcfg,
        },
        "paths": {
            "out_dir": str(out_dir.resolve()),
            "cache_dir": str((out_dir / ".wb_cache").resolve()),
        },
        "env": {
            "python": sys.version.split()[0],
            "platform": sys.platform,
            "cwd": str(Path.cwd()),
        },
    }

    write_json(manifest_path, manifest)
    print(f"[A] wrote manifest: {manifest_path}")
    print(f"[A] scope_count: {len(scope_list)}")
    return manifest_path


### сделать патч
#  # было:
# hashlib.sha1(s.encode("utf-8"), usedforsecurity=False).hexdigest()

# надо:
# hashlib.sha1(s.encode("utf-8")).hexdigest()

### "wb": {"dest": -1257786, "locale": "ru"}


# =========================
# Stage B: OWN FETCH (WB Necromancer v2)
# Depends on helpers from Stage A:
#   ensure_dir, write_json, read_json, utc_now_iso
# =========================

# from __future__ import annotations

import json
import re
import sys
import time
import math
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests

WB_CARD_V4 = "https://card.wb.ru/cards/v4/detail"
WB_CARD_V1 = "https://card.wb.ru/cards/v1/detail"


def _req_session() -> requests.Session:
    # English technical: stable headers, keep-alive, sane defaults
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Connection": "keep-alive",
    })
    return s


def wb_get_json(
    sess: requests.Session,
    url: str,
    params: Optional[Dict[str, Any]] = None,
    *,
    timeout: int = 30,
    retries: int = 3,
    backoff: float = 0.8,
) -> Tuple[Optional[dict], Dict[str, Any]]:
    """GET JSON with retry/backoff. Returns (json_or_none, info)."""
    last_err = None
    for attempt in range(retries + 1):
        try:
            r = sess.get(url, params=params, timeout=timeout)
            info = {"url": url, "params": params, "status": r.status_code}
            if r.status_code == 200:
                return r.json(), info
            if r.status_code in (429, 500, 502, 503, 504):
                time.sleep(backoff * (attempt + 1))
                continue
            # non-retryable
            return None, {**info, "error": f"http_{r.status_code}", "text_snip": r.text[:200]}
        except Exception as e:
            last_err = repr(e)
            time.sleep(backoff * (attempt + 1))
    return None, {"url": url, "params": params, "status": None, "error": last_err or "unknown"}


def fetch_card_v4(
    sess: requests.Session,
    nm_id: str,
    *,
    dest: int,
    locale: str,
    timeout: int,
) -> Tuple[Optional[dict], Dict[str, Any]]:
    params = {"appType": 1, "curr": "rub", "dest": dest, "locale": locale, "nm": nm_id}
    return wb_get_json(sess, WB_CARD_V4, params=params, timeout=timeout)


def fetch_card_v1(
    sess: requests.Session,
    nm_id: str,
    *,
    dest: int,
    timeout: int,
) -> Tuple[Optional[dict], Dict[str, Any]]:
    params = {"appType": 1, "curr": "rub", "dest": dest, "nm": nm_id}
    return wb_get_json(sess, WB_CARD_V1, params=params, timeout=timeout)


def nm_to_vol_part(nm_id: str) -> Tuple[int, int]:
    nm = int(nm_id)
    vol = nm // 100000
    part = nm // 1000
    return vol, part


def fetch_deep_card(
    sess: requests.Session,
    nm_id: str,
    *,
    locale: str = "ru",
    timeout: int = 30,
    max_baskets: int = 12,
) -> Tuple[Optional[dict], Dict[str, Any]]:
    """
    Try to fetch basket card.json.
    We brute-force basket hosts because mapping changes sometimes.
    """
    try:
        vol, part = nm_to_vol_part(nm_id)
    except Exception as e:
        return None, {"error": f"bad_nm_id:{e!r}"}

    path = f"/vol{vol}/part{part}/{nm_id}/info/{locale}/card.json"
    last_info: Dict[str, Any] = {}
    for i in range(1, max_baskets + 1):
        host = f"https://basket-{i:02d}.wb.ru"
        url = host + path
        js, info = wb_get_json(sess, url, params=None, timeout=timeout, retries=1, backoff=0.4)
        last_info = info
        if js is not None:
            info["basket"] = i
            return js, info
        # 404/403 are normal here, try next
        continue

    return None, {**last_info, "error": "deep_card_not_found"}


def extract_first_product(card_json: dict) -> Optional[dict]:
    try:
        data = card_json.get("data") or {}
        prods = data.get("products") or data.get("productsV2") or []
        if isinstance(prods, list) and prods:
            return prods[0]
    except Exception:
        return None
    return None


def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or (j.get("own") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done


def _price_to_rub(x: Any) -> Optional[float]:
    """
    WB часто отдаёт priceU/salePriceU в "копейках * 1" (или иногда *100),
    поэтому делаем простую эвристику. Нам нужна робастная оценка, не бухгалтерия.
    """
    try:
        if x is None:
            return None
        v = float(x)
        if not math.isfinite(v):
            return None
        # If value looks too large, treat as "minor units" and scale down.
        if v > 100000:  # >100k rub threshold
            return round(v / 100.0, 2)
        return round(v, 2)
    except Exception:
        return None


def normalize_own_record(
    nm_id: str,
    scope_row: dict,
    *,
    run_id: str,
    card_prod: Optional[dict],
    deep_json: Optional[dict],
    fetch_info: Dict[str, Any],
    deep_info: Optional[Dict[str, Any]],
) -> dict:
    vendor_code = str(scope_row.get("vendor_code") or "")
    scope_name = str(scope_row.get("name") or "")
    potential_qty = scope_row.get("potential_qty", None)

    own: Dict[str, Any] = {
        "nm_id": nm_id,
        "vendor_code": vendor_code,
        "scope_name": scope_name,
        "potential_qty": potential_qty,

        "imt_id": None,
        "subject_id": None,
        "subject_name": None,
        "brand": None,
        "name": None,

        "supplier_id": None,
        "supplier_name": None,

        "rating": None,
        "feedbacks": None,

        # For reports only (we ignore own price in decisions)
        "price_rub": None,
        "sale_price_rub": None,

        "pics": None,

        # Intent extraction inputs:
        "description": None,
        "options": None,
    }

    if card_prod:
        own["imt_id"] = card_prod.get("imtId") or card_prod.get("imt_id")
        own["subject_id"] = card_prod.get("subjId") or card_prod.get("subjectId")
        own["subject_name"] = card_prod.get("subjName") or card_prod.get("subjectName")

        own["brand"] = card_prod.get("brand")
        own["name"] = card_prod.get("name") or card_prod.get("title")

        own["supplier_id"] = card_prod.get("supplierId") or card_prod.get("sellerId")
        own["supplier_name"] = card_prod.get("supplier") or card_prod.get("seller")

        own["rating"] = card_prod.get("rating")
        own["feedbacks"] = card_prod.get("feedbacks")

        own["price_rub"] = _price_to_rub(card_prod.get("priceU") or card_prod.get("price"))
        own["sale_price_rub"] = _price_to_rub(card_prod.get("salePriceU") or card_prod.get("salePrice"))

        own["pics"] = card_prod.get("pics") or card_prod.get("picsCount") or None

    if deep_json and isinstance(deep_json, dict):
        desc = deep_json.get("description") or deep_json.get("desc")
        options = deep_json.get("options") or deep_json.get("characteristics") or deep_json.get("params")

        if desc:
            own["description"] = desc
        if options:
            own["options"] = options

        if not own["name"]:
            own["name"] = deep_json.get("name") or own["name"]
        if not own["brand"]:
            own["brand"] = deep_json.get("brand") or own["brand"]

    rec = {
        "meta": {
            "schema_version": "2.0",
            "run_id": run_id,
            "nm_id": nm_id,
            "vendor_code": vendor_code,
            "ts": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            "stage": "B",
        },
        "own": own,
        "fetch": {
            "card": fetch_info,
            "deep": deep_info,
        },
    }
    return rec


def stage_B_own_fetch(
    out_dir: Path,
    *,
    dest: int = -1257786,
    locale: str = "ru",
    timeout: int = 30,
    sleep_s: float = 0.4,
    deep_card: bool = True,
    resume: bool = False,
    strict: bool = False,
) -> Path:
    """
    Stage B: fetch own-card snapshots and write own_norm.jsonl.
    - Reads scope only from run_manifest.json.
    - Writes one JSONL row per SKU.
    - Stores raw snapshots in .wb_cache/own/{nm_id}/...
    """
    manifest_path = out_dir / "run_manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"run_manifest.json not found in {out_dir}. Run Stage A first.")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    scope_list = ((manifest.get("scope") or {}).get("sku_list") or [])

    if not run_id or not isinstance(scope_list, list) or not scope_list:
        raise ValueError("Bad manifest: missing run_id or scope.sku_list")

    out_path = out_dir / "own_norm.jsonl"
    err_path = out_dir / "own_errors.jsonl"

    done = read_jsonl_nm_ids(out_path) if resume else set()
    sess = _req_session()

    total = len(scope_list)
    for idx, row in enumerate(scope_list, 1):
        nm_id = str(row.get("nm_id") or "")
        if not nm_id:
            continue
        if nm_id in done:
            continue

        sku_cache_dir = out_dir / ".wb_cache" / "own" / nm_id
        ensure_dir(sku_cache_dir)

        card_v4_path = sku_cache_dir / "card_v4.json"
        card_v1_path = sku_cache_dir / "card_v1.json"
        deep_path = sku_cache_dir / "card_deep.json"

        card_json: Optional[dict] = None
        fetch_info: Dict[str, Any] = {}
        deep_json: Optional[dict] = None
        deep_info: Optional[Dict[str, Any]] = None

        try:
            # v4 first (richer)
            if card_v4_path.exists():
                card_json = read_json(card_v4_path)
                fetch_info = {"cached": True, "path": str(card_v4_path)}
            else:
                card_json, fetch_info = fetch_card_v4(sess, nm_id, dest=dest, locale=locale, timeout=timeout)
                fetch_info.update({"cached": False, "dest": dest, "locale": locale})
                if card_json is not None:
                    write_json(card_v4_path, card_json)

            # fallback to v1 if v4 empty
            if card_json is None:
                if card_v1_path.exists():
                    card_json = read_json(card_v1_path)
                    fetch_info = {"cached": True, "path": str(card_v1_path)}
                else:
                    card_json, fetch_info = fetch_card_v1(sess, nm_id, dest=dest, timeout=timeout)
                    fetch_info.update({"cached": False, "dest": dest})
                    if card_json is not None:
                        write_json(card_v1_path, card_json)

            prod = extract_first_product(card_json) if card_json else None

            # deep card (optional, best-effort)
            if deep_card:
                if deep_path.exists():
                    deep_json = read_json(deep_path)
                    deep_info = {"cached": True, "path": str(deep_path)}
                else:
                    deep_json, deep_info = fetch_deep_card(sess, nm_id, locale=locale, timeout=timeout)
                    if deep_json is not None:
                        write_json(deep_path, deep_json)

            rec = normalize_own_record(
                nm_id=nm_id,
                scope_row=row,
                run_id=run_id,
                card_prod=prod,
                deep_json=deep_json,
                fetch_info=fetch_info,
                deep_info=deep_info,
            )
            append_jsonl(out_path, rec)

            if idx % 10 == 0 or idx == total:
                print(f"[B] progress: {idx}/{total}", file=sys.stderr)

            time.sleep(max(0.0, float(sleep_s)))

        except Exception as e:
            err = {
                "meta": {
                    "schema_version": "2.0",
                    "run_id": run_id,
                    "nm_id": nm_id,
                    "vendor_code": str(row.get("vendor_code") or ""),
                    "ts": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
                    "stage": "B",
                },
                "error": repr(e),
                "fetch": {"card": fetch_info, "deep": deep_info},
            }
            append_jsonl(err_path, err)
            if strict:
                raise
            continue

    return out_path


# =========================
# Stage C: INTENT EXTRACT (WB Necromancer v2)
# Reads:  run_manifest.json, own_norm.jsonl
# Writes: intent.jsonl
# Network: NO
# =========================

# from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# Reuse helpers from earlier stages if they exist.
# If not yet centralized, keep these minimal local versions.
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _flatten_options(options: Any) -> str:
    """
    English technical:
    WB deep card options/characteristics vary wildly.
    We stringify values to enrich text for model detection.
    """
    if options is None:
        return ""
    try:
        if isinstance(options, dict):
            parts = []
            for k, v in options.items():
                parts.append(str(k))
                parts.append(str(v))
            return " ".join(parts)

        if isinstance(options, list):
            parts = []
            for it in options:
                if isinstance(it, dict):
                    for k in ("name", "param", "key"):
                        if k in it and it[k]:
                            parts.append(str(it[k]))
                    for k in ("value", "val", "text"):
                        if k in it and it[k]:
                            parts.append(str(it[k]))
                    if "values" in it and isinstance(it["values"], list):
                        parts.extend([str(x) for x in it["values"] if x is not None])
                else:
                    parts.append(str(it))
            return " ".join(parts)

        return str(options)
    except Exception:
        return ""


def _norm_text(s: str) -> str:
    s = (s or "").lower()
    s = s.replace("ё", "е")
    s = re.sub(r"[\u00A0\t\r\n]+", " ", s)
    s = re.sub(r"[^\w\+\- ]+", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def _detect_features(text: str) -> Dict[str, bool]:
    """
    Minimal feature flags for our v2 goals (TPU + pocket).
    """
    t = text
    has_tpu = any(w in t for w in ["tpu", "термополиуретан", "силикон", "силиконовый", "полиуретан"])
    has_pocket = any(w in t for w in [
        "карман", "картхолдер", "cardholder", "держатель карт",
        "для карт", "под карту", "для карты"
    ])
    has_magsafe = "magsafe" in t or "магсейф" in t or "магсейв" in t
    has_ring = any(w in t for w in ["кольцо", "ring holder", "держатель кольцо"])
    has_stand = any(w in t for w in ["подставка", "stand", "стенд"])
    return {
        "tpu": bool(has_tpu),
        "pocket": bool(has_pocket),
        "magsafe": bool(has_magsafe),
        "ring": bool(has_ring),
        "stand": bool(has_stand),
    }


def _extract_phone_model_candidates(raw_text: str) -> List[str]:
    """
    Heuristic extraction of phone model strings from RU/EN text.
    Returns candidates in 'pretty' form (not fully normalized).
    """
    t = _norm_text(raw_text)
    cands: List[str] = []

    # iPhone (iphone / айфон)
    for m in re.finditer(r"\b(?:iphone|айфон)\s*(\d{1,2})(?:\s*(pro|max|plus|mini))?(?:\s*(max))?\b", t):
        num = m.group(1)
        a = m.group(2) or ""
        b = m.group(3) or ""
        tail = " ".join([x for x in [a, b] if x]).strip()
        cand = f"iPhone {num}" + (f" {tail}" if tail else "")
        cands.append(cand.strip())

    # Samsung Galaxy A/S/M series (A52, S23 Ultra, M51 etc.)
    for m in re.finditer(r"\b(?:samsung\s*)?(?:galaxy\s*)?([asm]\s*\d{2,3})(?:\s*(fe|plus|ultra))?\b", t):
        base = re.sub(r"\s+", "", m.group(1)).upper()
        tail = (m.group(2) or "").upper()
        cand = f"Samsung Galaxy {base}" + (f" {tail}" if tail else "")
        cands.append(cand.strip())

    # Xiaomi / Redmi Note
    for m in re.finditer(r"\b(?:xiaomi\s*)?(?:redmi\s*)?(?:note\s*)?(\d{1,2})(?:\s*(pro\+|pro\s*\+|pro|max|plus|5g|4g))?\b", t):
        num = m.group(1)
        tail = (m.group(2) or "").replace(" ", "").upper()
        # avoid capturing random numbers; require context token nearby
        span = t[max(0, m.start()-12):m.end()+12]
        if any(k in span for k in ["redmi", "xiaomi", "note"]):
            cand = f"Redmi Note {num}" + (f" {tail}" if tail else "")
            cands.append(cand.strip())

    # POCO (X7, F5, M6 etc.)
    for m in re.finditer(r"\b(?:poco)\s*([xfm])\s*(\d{1,2})(?:\s*(pro|ultra|plus|5g|4g))?\b", t):
        series = m.group(1).upper()
        num = m.group(2)
        tail = (m.group(3) or "").upper()
        cand = f"POCO {series}{num}" + (f" {tail}" if tail else "")
        cands.append(cand.strip())

    # realme
    for m in re.finditer(r"\brealme\s*(\d{1,2})(?:\s*(pro\+|pro\s*\+|pro|plus|5g|4g))?\b", t):
        num = m.group(1)
        tail = (m.group(2) or "").replace(" ", "").upper()
        cand = f"realme {num}" + (f" {tail}" if tail else "")
        cands.append(cand.strip())

    # Honor / Huawei generic
    for m in re.finditer(r"\b(honor|huawei)\s*([a-z]*)\s*(\d{1,3}\w*)\b", t):
        brand = m.group(1).capitalize()
        series = (m.group(2) or "").upper()
        model = m.group(3).upper()
        cand = f"{brand} {series}{model}".strip()
        cands.append(cand)

    # Vivo / OPPO basic
    for m in re.finditer(r"\b(vivo|oppo)\s*([a-z]*)\s*(\d{1,3}\w*)\b", t):
        brand = m.group(1).upper()
        series = (m.group(2) or "").upper()
        model = m.group(3).upper()
        cand = f"{brand} {series}{model}".strip()
        cands.append(cand)

    # Dedupe preserve order
    seen = set()
    out: List[str] = []
    for c in cands:
        c2 = re.sub(r"\s{2,}", " ", c).strip()
        if not c2:
            continue
        k = c2.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(c2)
    return out


def _pick_best_model(cands: List[str], raw_text: str) -> Tuple[str, float]:
    """
    Pick best candidate by simple scoring:
    - longer (but not too long)
    - appears in title area (first 200 chars)
    - appears multiple times
    Returns (best, score). If none: ("", 0.0)
    """
    if not cands:
        return "", 0.0
    t = _norm_text(raw_text)
    head = t[:200]

    best = ""
    best_score = 0.0
    for c in cands:
        c_n = _norm_text(c)
        score = min(len(c_n), 40) / 40.0
        if c_n and c_n in head:
            score += 0.6
        freq = t.count(c_n) if c_n else 0
        score += min(freq, 3) * 0.25

        if score > best_score:
            best_score = score
            best = c
    return best, best_score


def _build_cluster_terms(phone_model: str, filters_cfg: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Build must/ban terms for each cluster.
    - phone cluster: must contain phone model tokens
    - type cluster: must contain phone model + TPU + pocket (as must_any groups)
    """
    pm = phone_model.strip()
    pm_terms: List[str] = []
    if pm:
        pm_terms = [x for x in re.split(r"\s+", _norm_text(pm)) if x]

    ban_default = list(filters_cfg.get("ban_terms_default") or [])
    tpu_terms = list(filters_cfg.get("tpu_terms") or ["tpu", "силикон", "силиконовый", "термополиуретан"])
    pocket_terms = list(filters_cfg.get("pocket_terms") or ["карман", "для карт", "под карту", "держатель карт", "cardholder"])

    return {
        "phone": {
            "must_terms": pm_terms,
            "ban_terms": ban_default,
            "must_any_groups": [
                {"name": "model_tokens", "any": pm_terms},
            ],
        },
        "type": {
            "must_terms": pm_terms + ["tpu", "карман"] if pm_terms else ["tpu", "карман"],
            "ban_terms": ban_default,
            "must_any_groups": [
                {"name": "model_tokens", "any": pm_terms},
                {"name": "tpu", "any": tpu_terms},
                {"name": "pocket", "any": pocket_terms},
            ],
        },
    }


def _karma_assess(rating: Any, feedbacks: Any, karma_cfg: Dict[str, Any]) -> Dict[str, Any]:
    """
    Determine if 'karma' is toxic enough to justify CLONE later.
    Stage C only sets the flag, stage L decides.
    """
    try:
        r = float(rating) if rating is not None else None
    except Exception:
        r = None

    try:
        f = int(feedbacks) if feedbacks is not None else 0
    except Exception:
        f = 0

    min_fb = int(karma_cfg.get("min_feedbacks_for_toxic", 20))
    toxic_rating_lt = float(karma_cfg.get("toxic_rating_lt", 4.2))
    very_toxic_lt = float(karma_cfg.get("very_toxic_rating_lt", 4.0))

    toxic = False
    level = "unknown"
    reasons: List[str] = []

    if r is None:
        level = "unknown"
        reasons.append("no_rating")
    else:
        level = "ok"
        if f >= min_fb and r < toxic_rating_lt:
            toxic = True
            level = "toxic"
            reasons.append(f"rating<{toxic_rating_lt} with feedbacks>={min_fb}")
        if f >= min_fb and r < very_toxic_lt:
            toxic = True
            level = "very_toxic"
            reasons.append(f"rating<{very_toxic_lt} with feedbacks>={min_fb}")

    return {
        "rating": r,
        "feedbacks": f,
        "toxic": toxic,
        "level": level,
        "reasons": reasons,
    }


def stage_C_intent_extract(
    out_dir: Path,
    *,
    resume: bool = False,
    karma_cfg: Optional[Dict[str, Any]] = None,
) -> Path:
    """
    Stage C: read own_norm.jsonl, produce intent.jsonl
    - rules-first extraction of phone_model and features
    - builds per-cluster must/ban terms
    """
    manifest_path = out_dir / "run_manifest.json"
    own_path = out_dir / "own_norm.jsonl"
    if not manifest_path.exists():
        raise FileNotFoundError(f"run_manifest.json not found in {out_dir}. Run Stage A first.")
    if not own_path.exists():
        raise FileNotFoundError(f"own_norm.jsonl not found in {out_dir}. Run Stage B first.")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    filters_cfg = ((manifest.get("config") or {}).get("filters") or {})
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    kc = {"min_feedbacks_for_toxic": 20, "toxic_rating_lt": 4.2, "very_toxic_rating_lt": 4.0}
    if isinstance(karma_cfg, dict):
        kc.update(karma_cfg)

    out_path = out_dir / "intent.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    processed = 0
    missing_models = 0

    with own_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            rec = json.loads(line)
            meta = rec.get("meta") or {}
            nm_id = str(meta.get("nm_id") or "")
            if not nm_id or nm_id in done:
                continue

            own = rec.get("own") or {}

            title = str(own.get("name") or "")
            scope_name = str(own.get("scope_name") or "")
            desc = str(own.get("description") or "")
            opts = _flatten_options(own.get("options"))

            raw_blob = " ".join([title, scope_name, desc, opts]).strip()
            norm_blob = _norm_text(raw_blob)

            features = _detect_features(norm_blob)

            cands = _extract_phone_model_candidates(raw_blob)
            best, score = _pick_best_model(cands, raw_blob)

            issues: List[str] = []
            if not best:
                missing_models += 1
                issues.append("MISSING_PHONE_MODEL")
            if not features.get("tpu", False):
                issues.append("NO_TPU_SIGNAL")
            if not features.get("pocket", False):
                issues.append("NO_POCKET_SIGNAL")

            cluster_terms = _build_cluster_terms(best, filters_cfg)
            karma = _karma_assess(own.get("rating"), own.get("feedbacks"), kc)

            intent = {
                "phone_model": {
                    "best": best,
                    "score": round(float(score), 3),
                    "candidates": cands[:10],
                },
                "case_type_guess": "phone_case",
                "feature_flags": features,
                "karma": karma,
                "clusters": [
                    {
                        "cluster": "phone",
                        "must_terms": cluster_terms["phone"]["must_terms"],
                        "ban_terms": cluster_terms["phone"]["ban_terms"],
                        "must_any_groups": cluster_terms["phone"]["must_any_groups"],
                        "query_seed": best or title or scope_name,
                    },
                    {
                        "cluster": "type",
                        "must_terms": cluster_terms["type"]["must_terms"],
                        "ban_terms": cluster_terms["type"]["ban_terms"],
                        "must_any_groups": cluster_terms["type"]["must_any_groups"],
                        "query_seed": (best + " tpu карман").strip() if best else (title or scope_name),
                    },
                ],
                "issues": issues,
            }

            out_rec = {
                "meta": {
                    "schema_version": "2.0",
                    "run_id": run_id,
                    "nm_id": nm_id,
                    "vendor_code": str(meta.get("vendor_code") or own.get("vendor_code") or ""),
                    "ts": _iso_utc(),
                    "stage": "C",
                },
                "intent": intent,
                "source": {
                    "title": title,
                    "scope_name": scope_name,
                    "has_description": bool(desc),
                    "has_options": bool(opts),
                },
            }

            append_jsonl(out_path, out_rec)
            processed += 1

    print(f"[C] processed: {processed}")
    if processed:
        print(f"[C] missing phone model: {missing_models} ({round(missing_models/processed*100, 1)}%)")
    return out_path


# =========================
# Stage D: QUERY BUILD (WB Necromancer v2)
# Reads:  run_manifest.json, intent.jsonl
# Writes: queries_raw.jsonl
# Network: NO (LLM enrichment is a hook; keep 0 until wired)
# =========================

# from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# --- minimal helpers (expected to exist in main file; keep local if not yet centralized) ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# --- query utils ---
def _norm(s: str) -> str:
    s = (s or "").lower().replace("ё", "е")
    s = re.sub(r"[\u00A0\t\r\n]+", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s

def _clean_query(q: str) -> str:
    q = _norm(q)
    q = re.sub(r"[\"'“”]+", "", q)
    q = re.sub(r"[^\w\-\+ ]+", " ", q, flags=re.UNICODE)
    q = re.sub(r"\s{2,}", " ", q).strip()
    return q

def _contains_any(text: str, terms: List[str]) -> bool:
    t = _norm(text)
    for x in terms:
        x = _norm(x)
        if x and x in t:
            return True
    return False

def _dedupe_keep_order(items: List[str]) -> List[str]:
    seen = set()
    out = []
    for it in items:
        k = _norm(it)
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(it)
    return out

def _extract_model_tokens(model_best: str) -> List[str]:
    # For soft enforcement; keep alnum tokens only.
    t = _clean_query(model_best)
    toks = [x for x in t.split() if x and len(x) >= 2]
    return toks


# --- rules-first query builders ---
def _build_rules_queries_phone(model: str) -> List[str]:
    if not model:
        return []
    m = model.strip()
    return [
        f"чехол {m}",
        f"чехол на {m}",
        f"чехол для {m}",
        f"чехол силиконовый {m}",
        f"чехол tpu {m}",
        f"накладка {m}",
        f"бампер {m}",
    ]

def _build_rules_queries_type(model: str) -> List[str]:
    if not model:
        return []
    m = model.strip()
    return [
        f"чехол tpu {m} карман",
        f"чехол tpu {m} для карт",
        f"чехол {m} силиконовый с карманом",
        f"чехол {m} картхолдер",
        f"чехол {m} с держателем карт",
        f"накладка tpu {m} карман для карт",
    ]


def _validate_query(
    q: str,
    *,
    cluster: str,
    model_tokens: List[str],
    must_any_groups: List[dict],
    ban_terms: List[str],
    min_len: int,
    max_len: int,
) -> Tuple[bool, List[str]]:
    reasons: List[str] = []
    qc = _clean_query(q)

    if not qc or len(qc) < min_len:
        return False, ["too_short"]
    if len(qc) > max_len:
        return False, ["too_long"]

    # Ban terms: hard drop
    for bt in ban_terms:
        bt_n = _norm(bt)
        if bt_n and bt_n in qc:
            return False, [f"ban_term:{bt_n}"]

    # Must-any groups: each group must match at least one term
    for g in must_any_groups or []:
        any_terms = g.get("any") or []
        if any_terms and not _contains_any(qc, any_terms):
            reasons.append(f"missing_group:{g.get('name','group')}")
    if reasons:
        return False, reasons

    # For phone cluster: soft model enforcement if tokens exist
    if cluster == "phone" and model_tokens:
        t = _norm(qc)
        present = sum(1 for tok in model_tokens if _norm(tok) in t)
        if present < max(1, len(model_tokens) // 2):
            return False, ["missing_model_tokens"]

    return True, []


def _take_top_n(qs: List[str], n: int) -> List[str]:
    return qs[: max(0, int(n))]


def stage_D_query_build(
    out_dir: Path,
    *,
    resume: bool = False,
    rules_per_cluster: int = 8,
    llm_extra_per_cluster: int = 0,   # keep 0 until LLM is wired with strict validation
    min_len: int = 8,
    max_len: int = 64,
) -> Path:
    """
    Stage D: build query packs for each SKU, per cluster (phone/type).
    Writes queries_raw.jsonl.
    Rules-first. LLM enrichment is a later hook.
    """
    manifest_path = out_dir / "run_manifest.json"
    intent_path = out_dir / "intent.jsonl"
    if not manifest_path.exists():
        raise FileNotFoundError(f"run_manifest.json not found in {out_dir}. Run Stage A first.")
    if not intent_path.exists():
        raise FileNotFoundError(f"intent.jsonl not found in {out_dir}. Run Stage C first.")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    out_path = out_dir / "queries_raw.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    processed = 0
    with intent_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            meta = rec.get("meta") or {}
            nm_id = str(meta.get("nm_id") or "")
            if not nm_id or nm_id in done:
                continue

            intent = rec.get("intent") or {}
            phone_model = ((intent.get("phone_model") or {}).get("best") or "").strip()

            clusters = intent.get("clusters") or []
            c_cfg: Dict[str, dict] = {c.get("cluster"): c for c in clusters if isinstance(c, dict)}

            packs: List[dict] = []
            for cluster in ("phone", "type"):
                cfg = c_cfg.get(cluster) or {}
                ban_terms = list(cfg.get("ban_terms") or [])
                must_any_groups = list(cfg.get("must_any_groups") or [])
                model_tokens = _extract_model_tokens(phone_model)

                rules_qs = _build_rules_queries_phone(phone_model) if cluster == "phone" else _build_rules_queries_type(phone_model)
                rules_qs = [_clean_query(x) for x in rules_qs]
                rules_qs = [x for x in rules_qs if x]

                valid_rules: List[str] = []
                dropped: List[dict] = []
                for q in rules_qs:
                    ok, why = _validate_query(
                        q,
                        cluster=cluster,
                        model_tokens=model_tokens,
                        must_any_groups=must_any_groups,
                        ban_terms=ban_terms,
                        min_len=min_len,
                        max_len=max_len,
                    )
                    if ok:
                        valid_rules.append(q)
                    else:
                        dropped.append({"q": q, "reasons": why})

                valid_rules = _dedupe_keep_order(valid_rules)
                valid_rules = _take_top_n(valid_rules, rules_per_cluster)

                llm_qs: List[str] = []
                # LLM enrichment hook (disabled by default):
                # - will be added later with provider config + strict re-validation
                if llm_extra_per_cluster and phone_model:
                    llm_qs = []

                all_qs = _dedupe_keep_order(valid_rules + llm_qs)

                packs.append({
                    "cluster": cluster,
                    "query_seed": cfg.get("query_seed") or phone_model or "",
                    "queries": all_qs,
                    "rules_count": len(valid_rules),
                    "llm_count": len(llm_qs),
                    "dropped_rules": dropped[:20],  # debug cap
                    "must_any_groups": must_any_groups,
                    "ban_terms": ban_terms,
                })

            out_rec = {
                "meta": {
                    "schema_version": "2.0",
                    "run_id": run_id,
                    "nm_id": nm_id,
                    "vendor_code": str(meta.get("vendor_code") or ""),
                    "ts": _iso_utc(),
                    "stage": "D",
                },
                "query_packs": packs,
                "notes": {
                    "rules_per_cluster": int(rules_per_cluster),
                    "llm_extra_per_cluster": int(llm_extra_per_cluster),
                    "min_len": int(min_len),
                    "max_len": int(max_len),
                    "phone_model": phone_model,
                },
            }

            append_jsonl(out_path, out_rec)
            processed += 1

    print(f"[D] processed: {processed}")
    return out_path

# =========================
# Stage E: SERP SNAPSHOT + VALIDATION (WB Necromancer v2)
# Reads:  run_manifest.json, queries_raw.jsonl
# Writes: queries_valid.jsonl
# Caches: .wb_cache/serp/{cluster}/{nm_id}/{sha1(query)}.json
# Network: WB
# =========================

# from __future__ import annotations

import json
import random
import re
import time
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests


DEFAULT_DESTS = [-1257786, -1216601, -115136, -421732, 123585595]
DEFAULT_SEARCH_HOSTS = ["u-search.wb.ru", "search.wb.ru"]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
}

# --- minimal helpers (expected to exist earlier; local copies are fine) ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def write_json(path: Path, data: Any) -> None:
    ensure_dir(path.parent)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    tmp.replace(path)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def _norm(s: str) -> str:
    s = (s or "").lower().replace("ё", "е")
    s = re.sub(r"[\u00A0\t\r\n]+", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s

def sha1_hex(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()

# --- WB search plumbing (mirrors wb_revive.py) ---
WB_SESSION = requests.Session()
WB_SESSION.headers.update(HEADERS)
WB_SESSION.trust_env = True
WB_SESSION.proxies = {}

def wb_search_v18_url(query: str, dest: int, *, page: int = 1, limit: int = 100, sort: str = "popular", host: str = "u-search.wb.ru") -> str:
    return (f"https://{host}/exactmatch/ru/common/v18/search"
            f"?appType=1&curr=rub&dest={dest}"
            f"&lang=ru&inheritFilters=false&suppressSpellcheck=false"
            f"&query={requests.utils.quote(query)}"
            f"&page={page}&resultset=catalog&sort={sort}&spp=30&limit={limit}")

def backoff_sleep(attempt: int, base: float = 0.35, cap: float = 6.0) -> None:
    time.sleep(min(cap, base * (2 ** attempt) + random.random() * 0.2))

def _get(url: str, timeout: int, retries: int = 3) -> Tuple[int, Optional[dict], str]:
    last = ""
    for a in range(retries):
        try:
            r = WB_SESSION.get(url, timeout=timeout)
            last = r.text[:2000]
            if r.status_code == 200:
                try:
                    return 200, r.json(), last
                except Exception:
                    return 200, None, last
            if r.status_code in (429, 500, 502, 503, 504):
                backoff_sleep(a)
                continue
            return r.status_code, None, last
        except requests.RequestException:
            backoff_sleep(a)
    return 0, None, last

def parse_search_items(js: dict) -> List[dict]:
    if not isinstance(js, dict):
        return []
    products = None
    if isinstance(js.get("products"), list):
        products = js.get("products")
    else:
        data = js.get("data")
        if isinstance(data, dict) and isinstance(data.get("products"), list):
            products = data.get("products")
    if not isinstance(products, list):
        return []

    out = []
    for idx, p in enumerate(products, start=1):
        if not isinstance(p, dict):
            continue
        nm = str(p.get("id") or p.get("nmId") or "").strip()
        if not nm.isdigit():
            continue
        out.append({
            "pos": idx,
            "nm_id": nm,
            "name": str(p.get("name") or ""),
            "brand": str(p.get("brand") or ""),
            "seller": str(p.get("supplier") or p.get("supplierName") or ""),
            "priceU": p.get("priceU") or p.get("price"),
            "salePriceU": p.get("salePriceU") or p.get("salePrice"),
            "rating": p.get("rating") if p.get("rating") is not None else p.get("reviewRating") or p.get("nmReviewRating"),
            "feedbacks": p.get("feedbacks") if p.get("feedbacks") is not None else p.get("nmFeedbacks"),
            "subjectId": p.get("subjectId") or p.get("subject"),
            "subjectName": str(p.get("subjectName") or p.get("subjName") or ""),
            "raw": p,
        })
    return out

def _extract_search_price_rub(item: dict) -> Optional[int]:
    raw = item.get("raw") if isinstance(item.get("raw"), dict) else None
    def to_int(x: Any) -> Optional[int]:
        try:
            if x is None:
                return None
            return int(float(x))
        except Exception:
            return None

    price_u = to_int(item.get("priceU"))
    sale_u = to_int(item.get("salePriceU"))
    use = sale_u if sale_u is not None else price_u
    if use is None:
        return None

    raw_has_u = isinstance(raw, dict) and (("priceU" in raw) or ("salePriceU" in raw))
    if raw_has_u:
        return int(max(0, use) // 100)
    if use >= 10000:
        return int(use // 100)
    return int(use)

def _query_tokens(q: str) -> List[str]:
    q = _norm(q)
    q = re.sub(r"[^a-zа-я0-9 ]+", " ", q)
    toks = [t for t in q.split() if len(t) >= 4]
    stop = {"для", "в", "на", "и", "с", "по", "как", "или", "без", "под", "все"}
    toks = [t for t in toks if t not in stop]
    return toks[:10]

def _relevance_score_token_overlap(query: str, items: List[dict]) -> int:
    """
    Old-school rel: count how many query tokens appear in top names (top12).
    Returns 0..10 approx. This matches wb_revive.py spirit.
    """
    if not items:
        return 0
    q = str(query).strip()
    if q.isdigit():
        qnm = q
        for it in items[:30]:
            if str(it.get("nm_id")) == qnm:
                return 100
        return 0
    toks = _query_tokens(q)
    if not toks:
        return 0
    topnames = " ".join(str(x.get("name") or "") for x in items[:12]).lower()
    topnames = topnames.replace("ё", "е")
    return sum(1 for t in toks if t in topnames)

def fetch_search_best(query: str, dests: List[int], timeout: int, limit: int, hosts: List[str]) -> Tuple[str, int, Optional[dict], str, int, int, str]:
    """
    Returns: (status, http_code, json, url, dest_used, rel_score, host_used)
    """
    best = None  # (score, count, code, js, url, dest, host)
    for host in hosts:
        for dest in dests:
            url = wb_search_v18_url(query, dest, page=1, limit=limit, host=host)
            code, js, _ = _get(url, timeout=timeout)
            if code != 200 or not js:
                continue
            items = parse_search_items(js)
            score = _relevance_score_token_overlap(query, items)
            cand = (score, len(items), code, js, url, dest, host)
            if best is None or cand[:2] > best[:2]:
                best = cand
            # early accept if looks non-empty and somewhat relevant
            if (score >= 1 or (str(query).isdigit() and score >= 100)) and len(items) >= min(10, limit):
                return "ok", code, js, url, dest, score, host
    if best:
        score, _, code, js, url, dest, host = best
        return "ok", code, js, url, dest, score, host
    return "not_found", 404, None, "", 0, 0, ""


# --- cluster-aware validation ---
CASE_INTENT_TERMS = ["чехол", "чехл", "case", "кейс", "наклад", "книжк", "бампер"]
POCKET_HINT_TERMS = ["карман", "картхолдер", "cardholder", "держатель карт", "для карт", "под карту", "слот", "отсек"]
TPU_HINT_TERMS = ["tpu", "силикон", "силиконовый", "термополиуретан", "полиуретан"]

def _contains_any(text: str, terms: List[str]) -> bool:
    t = _norm(text)
    for x in terms:
        x = _norm(x)
        if x and x in t:
            return True
    return False

def _pass_must_any_groups(text: str, must_any_groups: List[dict]) -> bool:
    t = _norm(text)
    for g in must_any_groups or []:
        any_terms = g.get("any") or []
        if any_terms and not _contains_any(t, [str(x) for x in any_terms]):
            return False
    return True

def _cluster_item_match(cluster: str, item_name: str) -> bool:
    """
    English technical:
    We only have item names at SERP stage.
    So cluster validation is name-based and intentionally conservative.
    """
    n = _norm(item_name)
    if not _contains_any(n, CASE_INTENT_TERMS):
        return False
    if cluster == "type":
        # Require *signals* for pocket and TPU in names in top listings (not perfect, but useful).
        if not _contains_any(n, POCKET_HINT_TERMS):
            return False
        if not _contains_any(n, TPU_HINT_TERMS):
            return False
    return True

def _serp_quality_metrics(cluster: str, items: List[dict], must_any_groups: List[dict]) -> Dict[str, Any]:
    """
    Compute pass counts over top50.
    """
    top = items[:50]
    if not top:
        return {"top_n": 0, "pass_n": 0, "pass_rate": 0.0, "case_rate": 0.0}

    names = [str(x.get("name") or "") for x in top]
    # group gating uses query intent tokens (model tokens, etc.)
    group_pass = [nm for nm in names if _pass_must_any_groups(nm, must_any_groups)]
    case_like = [nm for nm in names if _contains_any(nm, CASE_INTENT_TERMS)]

    if cluster == "type":
        cluster_pass = [nm for nm in group_pass if _cluster_item_match(cluster, nm)]
    else:
        # phone cluster: require case intent, plus must-any groups (model tokens)
        cluster_pass = [nm for nm in group_pass if _contains_any(nm, CASE_INTENT_TERMS)]

    top_n = len(names)
    pass_n = len(cluster_pass)
    return {
        "top_n": top_n,
        "pass_n": pass_n,
        "pass_rate": round(pass_n / max(1, top_n), 3),
        "case_rate": round(len(case_like) / max(1, top_n), 3),
    }

def _pick_valid_queries(per_query: List[dict], *, min_keep: int = 2, max_keep: int = 5) -> List[dict]:
    """
    Sort by (pass_rate, rel50, items_count) and keep top max_keep.
    """
    if not per_query:
        return []
    ranked = sorted(
        per_query,
        key=lambda x: (
            float((x.get("metrics") or {}).get("pass_rate") or 0.0),
            int(x.get("rel50") or 0),
            int(x.get("items_count") or 0),
        ),
        reverse=True,
    )
    keep = ranked[:max_keep]
    # if keep is too weak, allow fewer; stage will flag later anyway
    return keep[:max_keep]


def stage_E_serp_validate(
    out_dir: Path,
    *,
    timeout: int = 30,
    sleep_s: float = 0.4,
    search_limit: int = 100,
    resume: bool = False,
    min_keep_per_cluster: int = 2,
    max_keep_per_cluster: int = 5,
    min_pass_rate_phone: float = 0.12,
    min_pass_rate_type: float = 0.08,
) -> Path:
    """
    Stage E:
    - For each SKU and each cluster, fetch SERP for each query (best host+dest),
      cache raw JSON, compute rel50 and pass_rate on top50 names,
      then select 2-5 "valid" queries per cluster.
    """
    manifest_path = out_dir / "run_manifest.json"
    qraw_path = out_dir / "queries_raw.jsonl"
    if not manifest_path.exists():
        raise FileNotFoundError(f"run_manifest.json not found in {out_dir}. Run Stage A first.")
    if not qraw_path.exists():
        raise FileNotFoundError(f"queries_raw.jsonl not found in {out_dir}. Run Stage D first.")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    # Read WB config from manifest if present; else defaults
    wb_cfg = (manifest.get("config") or {}).get("wb") or (manifest.get("wb") or {})
    dests = list(wb_cfg.get("dests") or DEFAULT_DESTS)
    hosts = list(wb_cfg.get("search_hosts") or DEFAULT_SEARCH_HOSTS)

    out_path = out_dir / "queries_valid.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    processed = 0

    with qraw_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            rec = json.loads(line)
            meta = rec.get("meta") or {}
            nm_id = str(meta.get("nm_id") or "")
            if not nm_id or nm_id in done:
                continue

            packs = rec.get("query_packs") or []
            valid_packs: List[dict] = []
            all_notes: List[str] = []

            for pack in packs:
                cluster = str(pack.get("cluster") or "").strip()
                if cluster not in ("phone", "type"):
                    continue

                queries = [str(x) for x in (pack.get("queries") or []) if str(x).strip()]
                must_any_groups = list(pack.get("must_any_groups") or [])
                ban_terms = list(pack.get("ban_terms") or [])

                per_query_rows: List[dict] = []

                for q in queries:
                    qn = _norm(q)
                    # trivial ban guard (query-level)
                    if any(_norm(bt) in qn for bt in ban_terms if bt):
                        per_query_rows.append({
                            "query": q,
                            "status": "dropped_ban_term",
                            "rel50": 0,
                            "items_count": 0,
                            "metrics": {"top_n": 0, "pass_n": 0, "pass_rate": 0.0, "case_rate": 0.0},
                            "cache_path": None,
                        })
                        continue

                    status, code, js, url, dest_used, rel50, host_used = fetch_search_best(
                        q, dests=dests, timeout=timeout, limit=search_limit, hosts=hosts
                    )
                    if status != "ok" or not js:
                        per_query_rows.append({
                            "query": q,
                            "status": "not_found",
                            "http": code,
                            "rel50": 0,
                            "items_count": 0,
                            "metrics": {"top_n": 0, "pass_n": 0, "pass_rate": 0.0, "case_rate": 0.0},
                            "cache_path": None,
                            "fetch": {"url": url, "dest": dest_used, "host": host_used},
                        })
                        time.sleep(max(0.0, float(sleep_s)))
                        continue

                    items = parse_search_items(js)
                    metrics = _serp_quality_metrics(cluster, items, must_any_groups)

                    cache_dir = out_dir / ".wb_cache" / "serp" / cluster / nm_id
                    ensure_dir(cache_dir)
                    cache_path = cache_dir / f"{sha1_hex(qn)}.json"
                    if not cache_path.exists():
                        write_json(cache_path, js)

                    prices = [p for p in (_extract_search_price_rub(it) for it in items[:100]) if p is not None]
                    price_med = None
                    if prices:
                        ps = sorted(prices)
                        price_med = ps[len(ps) // 2]

                    per_query_rows.append({
                        "query": q,
                        "status": "ok",
                        "http": code,
                        "rel50": int(rel50),
                        "items_count": len(items),
                        "metrics": metrics,
                        "price_median_rub_top100": price_med,
                        "cache_path": str(cache_path),
                        "fetch": {"url": url, "dest": dest_used, "host": host_used},
                    })

                    time.sleep(max(0.0, float(sleep_s)))

                # Now select valid queries
                kept = _pick_valid_queries(per_query_rows, min_keep=min_keep_per_cluster, max_keep=max_keep_per_cluster)

                # Apply minimal pass-rate floor to avoid garbage "valid"
                floor = min_pass_rate_type if cluster == "type" else min_pass_rate_phone
                kept2 = []
                for row in kept:
                    pr = float((row.get("metrics") or {}).get("pass_rate") or 0.0)
                    if row.get("status") == "ok" and pr >= floor:
                        kept2.append(row)

                if len(kept2) < min_keep_per_cluster:
                    all_notes.append(f"LOW_CONF_{cluster}: kept={len(kept2)}<min_keep={min_keep_per_cluster}")

                valid_packs.append({
                    "cluster": cluster,
                    "selected": kept2,
                    "all_results": per_query_rows,  # keep for debug; can trim later
                })

            out_rec = {
                "meta": {
                    "schema_version": "2.0",
                    "run_id": run_id,
                    "nm_id": nm_id,
                    "vendor_code": str(meta.get("vendor_code") or ""),
                    "ts": _iso_utc(),
                    "stage": "E",
                },
                "valid_query_packs": valid_packs,
                "notes": all_notes,
                "wb": {"dests": dests, "search_hosts": hosts, "search_limit": int(search_limit)},
            }

            append_jsonl(out_path, out_rec)
            processed += 1

    print(f"[E] processed: {processed}")
    return out_path

# =========================
# Stage F: COMPETITOR POOL (WB Necromancer v2)
# Reads:  run_manifest.json, queries_valid.jsonl + SERP cache files
# Writes: competitor_pool.jsonl
# Network: NO (LOCAL only)
# =========================

# from __future__ import annotations

import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# --- minimal helpers (expected to exist earlier; local copies ok for now) ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def _norm(s: str) -> str:
    s = (s or "").lower().replace("ё", "е")
    s = re.sub(r"[\u00A0\t\r\n]+", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s

def _contains_any(text: str, terms: List[str]) -> bool:
    t = _norm(text)
    for x in terms:
        x = _norm(str(x))
        if x and x in t:
            return True
    return False

def _pass_must_any_groups(text: str, must_any_groups: List[dict]) -> bool:
    t = _norm(text)
    for g in must_any_groups or []:
        any_terms = g.get("any") or []
        if any_terms and not _contains_any(t, [str(x) for x in any_terms]):
            return False
    return True


# --- cluster intent signals (same spirit as Stage E) ---
CASE_INTENT_TERMS = ["чехол", "чехл", "case", "кейс", "наклад", "книжк", "бампер"]
POCKET_HINT_TERMS = ["карман", "картхолдер", "cardholder", "держатель карт", "для карт", "под карту", "слот", "отсек"]
TPU_HINT_TERMS = ["tpu", "силикон", "силиконовый", "термополиуретан", "полиуретан"]

def _cluster_item_match(cluster: str, item_name: str) -> bool:
    n = _norm(item_name)
    if not _contains_any(n, CASE_INTENT_TERMS):
        return False
    if cluster == "type":
        if not _contains_any(n, POCKET_HINT_TERMS):
            return False
        if not _contains_any(n, TPU_HINT_TERMS):
            return False
    return True


# --- parse SERP cache JSON (v18 search response variants) ---
def parse_search_items(js: dict) -> List[dict]:
    """
    English technical:
    Accepts either {products:[...]} or {data:{products:[...]}}.
    Extracts best-effort imtId/supplierId if present; otherwise keeps nm_id uniqueness.
    """
    if not isinstance(js, dict):
        return []
    products = None
    if isinstance(js.get("products"), list):
        products = js.get("products")
    else:
        data = js.get("data")
        if isinstance(data, dict) and isinstance(data.get("products"), list):
            products = data.get("products")
    if not isinstance(products, list):
        return []

    out = []
    for idx, p in enumerate(products, start=1):
        if not isinstance(p, dict):
            continue
        nm = str(p.get("id") or p.get("nmId") or "").strip()
        if not nm.isdigit():
            continue

        # Best-effort keys (WB responses vary)
        imt = p.get("imtId") or p.get("imt_id") or p.get("root") or p.get("rootId")
        try:
            imt = int(imt) if imt is not None else None
        except Exception:
            imt = None

        seller_id = p.get("supplierId") or p.get("sellerId") or p.get("supplier_id")
        try:
            seller_id = int(seller_id) if seller_id is not None else None
        except Exception:
            seller_id = None

        seller_name = str(p.get("supplier") or p.get("supplierName") or p.get("seller") or "")
        name = str(p.get("name") or "")
        brand = str(p.get("brand") or "")
        subject_id = p.get("subjectId") or p.get("subject")
        subject_name = str(p.get("subjectName") or p.get("subjName") or "")

        rating = p.get("rating")
        if rating is None:
            rating = p.get("reviewRating") or p.get("nmReviewRating")
        feedbacks = p.get("feedbacks")
        if feedbacks is None:
            feedbacks = p.get("nmFeedbacks")

        # priceU/salePriceU often in minor units; we store raw here (structure stage will recompute later)
        price_u = p.get("priceU") or p.get("price")
        sale_u = p.get("salePriceU") or p.get("salePrice")

        out.append({
            "pos": idx,
            "nm_id": nm,
            "imt_id": imt,
            "seller_id": seller_id,
            "seller_name": seller_name,
            "name": name,
            "brand": brand,
            "subject_id": subject_id,
            "subject_name": subject_name,
            "rating": rating,
            "feedbacks": feedbacks,
            "priceU": price_u,
            "salePriceU": sale_u,
        })
    return out


def _seller_key(item: dict) -> str:
    sid = item.get("seller_id")
    if sid is not None:
        return f"id:{sid}"
    sn = _norm(item.get("seller_name") or "")
    return f"name:{sn}" if sn else "unknown"


def _cand_key(item: dict) -> str:
    # Dedup by imt_id if present else by nm_id
    imt = item.get("imt_id")
    if imt is not None:
        return f"imt:{imt}"
    return f"nm:{item.get('nm_id')}"


def _safe_float(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        v = float(x)
        if not math.isfinite(v):
            return None
        return v
    except Exception:
        return None

def _safe_int(x: Any) -> Optional[int]:
    try:
        if x is None:
            return None
        return int(float(x))
    except Exception:
        return None


def _rank_candidates(cands: List[dict]) -> List[dict]:
    """
    Sort by:
    - intent_match_count (desc)
    - appearances (desc)
    - best_pos (asc)
    - rating (desc)
    - feedbacks (desc)
    """
    def key(c: dict):
        return (
            int(c.get("intent_match_count") or 0),
            int(c.get("appearances") or 0),
            -int(c.get("best_pos") or 10**9),  # inverted later by reverse=False trick? easier keep separate
        )
    # We'll do a clearer sort manually:
    return sorted(
        cands,
        key=lambda c: (
            -(int(c.get("intent_match_count") or 0)),
            -(int(c.get("appearances") or 0)),
            int(c.get("best_pos") or 10**9),
            -(int((c.get("rating") or 0) * 100)),
            -(int(c.get("feedbacks") or 0)),
        ),
    )


def _select_diverse(
    ranked: List[dict],
    *,
    take: int,
    max_per_seller: int,
    already: Optional[set] = None,
) -> List[dict]:
    already = already or set()
    out: List[dict] = []
    seller_counts: Dict[str, int] = {}

    for c in ranked:
        k = c.get("key")
        if not k or k in already:
            continue
        sk = c.get("seller_key") or "unknown"
        if seller_counts.get(sk, 0) >= max_per_seller:
            continue
        out.append(c)
        already.add(k)
        seller_counts[sk] = seller_counts.get(sk, 0) + 1
        if len(out) >= take:
            break
    return out


def stage_F_competitor_pool(
    out_dir: Path,
    *,
    resume: bool = False,
    top_k_per_query: int = 60,
    leaders_take: int = 18,
    closest_take: int = 18,
    max_total_per_cluster: int = 35,
    max_per_seller: int = 3,
) -> Path:
    """
    Stage F:
    - Build competitor pool per SKU per cluster from selected SERP caches.
    - Dedup by imt_id if present, else nm_id.
    - Diversify by seller.
    Writes competitor_pool.jsonl.
    """
    manifest_path = out_dir / "run_manifest.json"
    qvalid_path = out_dir / "queries_valid.jsonl"

    if not manifest_path.exists():
        raise FileNotFoundError(f"run_manifest.json not found in {out_dir}. Run Stage A first.")
    if not qvalid_path.exists():
        raise FileNotFoundError(f"queries_valid.jsonl not found in {out_dir}. Run Stage E first.")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    out_path = out_dir / "competitor_pool.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    processed = 0

    with qvalid_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            rec = json.loads(line)
            meta = rec.get("meta") or {}
            nm_id = str(meta.get("nm_id") or "")
            if not nm_id or nm_id in done:
                continue

            packs = rec.get("valid_query_packs") or []
            out_clusters: List[dict] = []
            notes: List[str] = list(rec.get("notes") or [])

            for pack in packs:
                cluster = str(pack.get("cluster") or "").strip()
                if cluster not in ("phone", "type"):
                    continue

                selected = pack.get("selected") or []
                # must_any_groups are stored per selected row inside Stage D packs; Stage E kept them in pack? Not guaranteed.
                # Stage E stored must_any_groups per pack only in queries_raw. Here we use best-effort:
                # - If Stage E stored must_any_groups inside rows, we can read; else no group gating.
                # We'll accept a pack-level hint if present.
                must_any_groups = list(pack.get("must_any_groups") or [])

                # Aggregate candidates from SERP caches
                agg: Dict[str, dict] = {}
                missing_cache = 0
                imt_missing = 0

                for row in selected:
                    cache_path = row.get("cache_path")
                    if not cache_path:
                        missing_cache += 1
                        continue
                    cp = Path(str(cache_path))
                    if not cp.exists():
                        missing_cache += 1
                        continue

                    js = read_json(cp)
                    items = parse_search_items(js)[: max(1, int(top_k_per_query))]

                    for it in items:
                        name = str(it.get("name") or "")
                        # Apply group gating if we have it, otherwise only case/type match
                        group_ok = _pass_must_any_groups(name, must_any_groups) if must_any_groups else True
                        intent_ok = group_ok and _cluster_item_match(cluster, name)

                        ck = _cand_key(it)
                        sk = _seller_key(it)

                        if it.get("imt_id") is None:
                            imt_missing += 1

                        if ck not in agg:
                            agg[ck] = {
                                "key": ck,
                                "nm_id": it.get("nm_id"),
                                "imt_id": it.get("imt_id"),
                                "seller_id": it.get("seller_id"),
                                "seller_name": it.get("seller_name"),
                                "seller_key": sk,
                                "name": name,
                                "brand": it.get("brand"),
                                "subject_id": it.get("subject_id"),
                                "subject_name": it.get("subject_name"),
                                "rating": _safe_float(it.get("rating")),
                                "feedbacks": _safe_int(it.get("feedbacks")),
                                "best_pos": int(it.get("pos") or 10**9),
                                "avg_pos_sum": int(it.get("pos") or 0),
                                "avg_pos_n": 1,
                                "appearances": 1,
                                "intent_match_count": 1 if intent_ok else 0,
                                "seen_in_queries": 1,
                            }
                        else:
                            c = agg[ck]
                            pos = int(it.get("pos") or 10**9)
                            c["best_pos"] = min(int(c.get("best_pos") or 10**9), pos)
                            c["avg_pos_sum"] = int(c.get("avg_pos_sum") or 0) + pos
                            c["avg_pos_n"] = int(c.get("avg_pos_n") or 0) + 1
                            c["appearances"] = int(c.get("appearances") or 0) + 1
                            c["intent_match_count"] = int(c.get("intent_match_count") or 0) + (1 if intent_ok else 0)
                            c["seen_in_queries"] = int(c.get("seen_in_queries") or 0) + 1

                cands = list(agg.values())
                for c in cands:
                    n = max(1, int(c.get("avg_pos_n") or 1))
                    c["avg_pos"] = round(int(c.get("avg_pos_sum") or 0) / n, 2)

                if not cands:
                    notes.append(f"EMPTY_POOL_{cluster}")
                    out_clusters.append({
                        "cluster": cluster,
                        "candidates": [],
                        "selected_nm_ids": [],
                        "stats": {"pool_size": 0, "unique_sellers": 0, "missing_cache": missing_cache},
                        "warnings": ["no_candidates_from_serp"],
                    })
                    continue

                ranked = _rank_candidates(cands)

                # Leaders: prioritize by best_pos from ranked (already uses best_pos)
                leaders = _select_diverse(
                    ranked,
                    take=leaders_take,
                    max_per_seller=max_per_seller,
                    already=set(),
                )

                # Closest: prioritize high intent_match_count; ranked already starts with that.
                already_keys = set(x["key"] for x in leaders if x.get("key"))
                closest = _select_diverse(
                    ranked,
                    take=closest_take,
                    max_per_seller=max_per_seller,
                    already=already_keys,
                )

                selected_all = leaders + closest
                # Cap total
                selected_all = selected_all[:max_total_per_cluster]

                selected_nm_ids = []
                for c in selected_all:
                    nm = str(c.get("nm_id") or "")
                    if nm and nm.isdigit():
                        selected_nm_ids.append(nm)

                selected_nm_ids = list(dict.fromkeys(selected_nm_ids))  # dedupe keep order

                unique_sellers = len(set((c.get("seller_key") or "unknown") for c in cands))
                imt_missing_rate = round(imt_missing / max(1, len(cands)), 3)

                warn: List[str] = []
                if missing_cache:
                    warn.append(f"missing_cache:{missing_cache}")
                if imt_missing_rate > 0.5:
                    warn.append("imt_id_often_missing_in_serp")  # final dedupe will happen after lite fetch

                out_clusters.append({
                    "cluster": cluster,
                    "candidates": ranked[:200],  # debug cap, still useful for audits
                    "selected_nm_ids": selected_nm_ids,
                    "stats": {
                        "pool_size": len(cands),
                        "selected_count": len(selected_nm_ids),
                        "unique_sellers": unique_sellers,
                        "imt_missing_rate": imt_missing_rate,
                        "missing_cache": missing_cache,
                    },
                    "warnings": warn,
                })

            out_rec = {
                "meta": {
                    "schema_version": "2.0",
                    "run_id": run_id,
                    "nm_id": nm_id,
                    "vendor_code": str(meta.get("vendor_code") or ""),
                    "ts": _iso_utc(),
                    "stage": "F",
                },
                "clusters": out_clusters,
                "notes": notes,
                "selection_policy": {
                    "top_k_per_query": int(top_k_per_query),
                    "leaders_take": int(leaders_take),
                    "closest_take": int(closest_take),
                    "max_total_per_cluster": int(max_total_per_cluster),
                    "max_per_seller": int(max_per_seller),
                    "dedupe_key": "imt_id if present else nm_id",
                },
            }

            append_jsonl(out_path, out_rec)
            processed += 1

    print(f"[F] processed: {processed}")
    return out_path

# =========================
# Stage G: LITE FETCH (WB Necromancer v2)
# Reads:  run_manifest.json, competitor_pool.jsonl
# Writes: competitor_lite.jsonl (+ competitor_lite_errors.jsonl)
# Caches: .wb_cache/comp_lite/{nm_id}.json
# Network: WB
# =========================

# from __future__ import annotations

import json
import sys
import time
import random
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests

WB_CARD_V4 = "https://card.wb.ru/cards/v4/detail"
WB_CARD_V1 = "https://card.wb.ru/cards/v1/detail"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
}

# --- minimal helpers (same vibe as wb_revive.py) ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def write_json(path: Path, data: Any) -> None:
    ensure_dir(path.parent)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    tmp.replace(path)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# --- WB client ---
def _req_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    s.trust_env = True  # let user run behind VPN/proxy if configured
    return s

def _backoff_sleep(attempt: int, base: float = 0.4, cap: float = 6.0) -> None:
    time.sleep(min(cap, base * (2 ** attempt) + random.random() * 0.25))

def wb_get_json(
    sess: requests.Session,
    url: str,
    params: Dict[str, Any],
    *,
    timeout: int,
    retries: int = 3
) -> Tuple[Optional[dict], Dict[str, Any]]:
    last_err = ""
    for a in range(retries + 1):
        try:
            r = sess.get(url, params=params, timeout=timeout)
            info = {"url": url, "params": params, "status": r.status_code}
            if r.status_code == 200:
                try:
                    return r.json(), info
                except Exception as e:
                    return None, {**info, "error": f"bad_json:{e!r}", "text_snip": r.text[:200]}
            if r.status_code in (429, 500, 502, 503, 504):
                _backoff_sleep(a)
                continue
            return None, {**info, "error": f"http_{r.status_code}", "text_snip": r.text[:200]}
        except Exception as e:
            last_err = repr(e)
            _backoff_sleep(a)
    return None, {"url": url, "params": params, "status": None, "error": last_err or "unknown"}

def fetch_card_v4(sess: requests.Session, nm_id: str, *, dest: int, locale: str, timeout: int) -> Tuple[Optional[dict], Dict[str, Any]]:
    params = {"appType": 1, "curr": "rub", "dest": dest, "locale": locale, "nm": nm_id}
    return wb_get_json(sess, WB_CARD_V4, params=params, timeout=timeout)

def fetch_card_v1(sess: requests.Session, nm_id: str, *, dest: int, timeout: int) -> Tuple[Optional[dict], Dict[str, Any]]:
    params = {"appType": 1, "curr": "rub", "dest": dest, "nm": nm_id}
    return wb_get_json(sess, WB_CARD_V1, params=params, timeout=timeout)

def extract_product(card_json: dict, nm_id: str) -> Optional[dict]:
    if not isinstance(card_json, dict):
        return None

    data = card_json.get("data")
    if isinstance(data, dict) and isinstance(data.get("products"), list):
        prods = data.get("products") or []
    elif isinstance(card_json.get("products"), list):
        prods = card_json.get("products") or []
    else:
        prods = []

    if not prods:
        return None

    # If multiple, try to match exact nm_id
    for p in prods:
        if isinstance(p, dict) and str(p.get("id") or p.get("nmId") or "") == str(nm_id):
            return p

    return prods[0] if isinstance(prods[0], dict) else None

def _price_to_rub_from_product(prod: dict) -> Tuple[Optional[float], Optional[float]]:
    def to_rub(v: Any) -> Optional[float]:
        try:
            if v is None:
                return None
            x = float(v)
            if x != x:
                return None
            # Heuristic: WB priceU usually in kopecks
            if x > 10000:
                return round(x / 100.0, 2)
            return round(x, 2)
        except Exception:
            return None

    price = to_rub(prod.get("priceU") or prod.get("price"))
    sale = to_rub(prod.get("salePriceU") or prod.get("salePrice"))
    return price, sale

def normalize_comp_lite(nm_id: str, prod: dict) -> dict:
    price_rub, sale_rub = _price_to_rub_from_product(prod or {})
    return {
        "nm_id": str(nm_id),
        "imt_id": prod.get("imtId") or prod.get("imt_id"),
        "subject_id": prod.get("subjectId") or prod.get("subjId") or prod.get("subject"),
        "subject_name": prod.get("subjectName") or prod.get("subjName") or "",
        "brand": prod.get("brand") or "",
        "name": prod.get("name") or prod.get("title") or "",
        "seller_id": prod.get("supplierId") or prod.get("sellerId") or None,
        "seller_name": prod.get("supplier") or prod.get("supplierName") or prod.get("seller") or "",
        "rating": prod.get("rating"),
        "feedbacks": prod.get("feedbacks"),
        "price_rub": price_rub,
        "sale_price_rub": sale_rub,
        # Stock proxy (best-effort; field names vary)
        "total_quantity": prod.get("totalQuantity") or prod.get("quantity") or prod.get("qty") or None,
        "pics": prod.get("pics") or prod.get("picsCount") or None,
    }


def stage_G_competitor_lite_fetch(
    out_dir: Path,
    *,
    dest: int = -1257786,
    locale: str = "ru",
    timeout: int = 30,
    sleep_s: float = 0.35,
    resume: bool = False,
    strict: bool = False,
) -> Path:
    """
    Stage G (WB):
    - Reads competitor_pool.jsonl (selected_nm_ids per cluster).
    - Fetches lite card data for each selected competitor nm_id (cached).
    - Writes competitor_lite.jsonl with per-cluster lists + merged list.
    """
    manifest_path = out_dir / "run_manifest.json"
    pool_path = out_dir / "competitor_pool.jsonl"
    if not manifest_path.exists():
        raise FileNotFoundError(f"run_manifest.json not found in {out_dir}. Run Stage A first.")
    if not pool_path.exists():
        raise FileNotFoundError(f"competitor_pool.jsonl not found in {out_dir}. Run Stage F first.")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    # Manifest override (menu/cli like wb_revive.py)
    wb_cfg = (manifest.get("config") or {}).get("wb") or (manifest.get("wb") or {})
    dest = int(wb_cfg.get("dest") or dest)
    locale = str(wb_cfg.get("locale") or locale)

    out_path = out_dir / "competitor_lite.jsonl"
    err_path = out_dir / "competitor_lite_errors.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    sess = _req_session()

    processed = 0
    with pool_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            rec = json.loads(line)
            meta = rec.get("meta") or {}
            own_nm = str(meta.get("nm_id") or "")
            if not own_nm or own_nm in done:
                continue

            clusters = rec.get("clusters") or []

            # Collect selected competitors per cluster
            selected_map: Dict[str, List[str]] = {}
            for c in clusters:
                cl = str(c.get("cluster") or "")
                if cl not in ("phone", "type"):
                    continue
                ids = [str(x) for x in (c.get("selected_nm_ids") or []) if str(x).isdigit()]
                selected_map[cl] = ids

            all_selected: List[str] = []
            for ids in selected_map.values():
                all_selected.extend(ids)

            # Dedupe keep order
            all_selected = list(dict.fromkeys(all_selected))

            cache_base = out_dir / ".wb_cache" / "comp_lite"
            ensure_dir(cache_base)

            fetched: Dict[str, dict] = {}
            fetch_stats = {"selected_total": len(all_selected), "fetched": 0, "cached": 0, "missing": 0, "errors": 0}

            for cid in all_selected:
                cache_path = cache_base / f"{cid}.json"
                card_json = None
                info: Dict[str, Any] = {}

                try:
                    if cache_path.exists():
                        card_json = read_json(cache_path)
                        fetch_stats["cached"] += 1
                        info = {"cached": True, "path": str(cache_path)}
                    else:
                        card_json, info = fetch_card_v4(sess, cid, dest=dest, locale=locale, timeout=timeout)
                        info.update({"cached": False, "dest": dest, "locale": locale})

                        if card_json is None:
                            # Fallback v1
                            card_json, info2 = fetch_card_v1(sess, cid, dest=dest, timeout=timeout)
                            info = {**info, "fallback_v1": info2}

                        if card_json is not None:
                            write_json(cache_path, card_json)

                    if card_json is None:
                        fetch_stats["missing"] += 1
                        fetched[cid] = {"nm_id": cid, "missing": True, "fetch": info}
                    else:
                        prod = extract_product(card_json, cid)
                        if not prod:
                            fetch_stats["missing"] += 1
                            fetched[cid] = {"nm_id": cid, "missing": True, "fetch": info}
                        else:
                            lite = normalize_comp_lite(cid, prod)
                            lite["fetch"] = {"cached": bool(info.get("cached")), "cache_path": str(cache_path)}
                            fetched[cid] = lite
                            fetch_stats["fetched"] += 1

                except Exception as e:
                    fetch_stats["errors"] += 1
                    append_jsonl(err_path, {
                        "meta": {
                            "schema_version": "2.0",
                            "run_id": run_id,
                            "nm_id": own_nm,
                            "ts": _iso_utc(),
                            "stage": "G",
                        },
                        "competitor_nm_id": cid,
                        "error": repr(e),
                    })
                    if strict:
                        raise
                finally:
                    time.sleep(max(0.0, float(sleep_s)))

            # Build merged list with cluster membership
            merged: List[dict] = []
            for cid in all_selected:
                item = fetched.get(cid) or {"nm_id": cid, "missing": True}
                membership = [cl for cl, ids in selected_map.items() if cid in ids]
                it2 = dict(item)
                it2["clusters"] = membership
                merged.append(it2)

            by_cluster: Dict[str, List[dict]] = {}
            for cl in ("phone", "type"):
                ids = selected_map.get(cl) or []
                by_cluster[cl] = [
                    next((x for x in merged if x.get("nm_id") == cid), {"nm_id": cid, "missing": True, "clusters": [cl]})
                    for cid in ids
                ]

            imt_ids = [x.get("imt_id") for x in merged if isinstance(x, dict) and x.get("imt_id") is not None]
            unique_imt = len(set(imt_ids)) if imt_ids else 0

            append_jsonl(out_path, {
                "meta": {
                    "schema_version": "2.0",
                    "run_id": run_id,
                    "nm_id": own_nm,
                    "vendor_code": str(meta.get("vendor_code") or ""),
                    "ts": _iso_utc(),
                    "stage": "G",
                },
                "competitors": {
                    "selected_nm_ids": selected_map,
                    "by_cluster": by_cluster,
                    "merged": merged,
                },
                "stats": {
                    **fetch_stats,
                    "unique_imt_ids": unique_imt,
                },
                "wb": {"dest": dest, "locale": locale, "timeout": int(timeout)},
            })

            processed += 1
            if processed % 10 == 0:
                print(f"[G] processed: {processed}", file=sys.stderr)

    print(f"[G] processed: {processed}")
    return out_path

# =========================
# Stage H: RELEVANCE FILTER (WB Necromancer v2)
# Reads:  run_manifest.json, intent.jsonl, competitor_lite.jsonl
# Writes: relevance.jsonl
# Network: NO (LOCAL only)
# =========================

# from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# --- minimal helpers ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def _norm(s: str) -> str:
    s = (s or "").lower().replace("ё", "е")
    s = re.sub(r"[\u00A0\t\r\n]+", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


# --- intent signals ---
CASE_INTENT_TERMS = ["чехол", "чехл", "case", "кейс", "наклад", "книжк", "бампер"]
POCKET_HINT_TERMS = ["карман", "картхолдер", "cardholder", "держатель карт", "для карт", "под карту", "слот", "отсек"]
TPU_HINT_TERMS = ["tpu", "силикон", "силиконовый", "термополиуретан", "полиуретан"]

def _contains_any(text: str, terms: List[str]) -> bool:
    t = _norm(text)
    for x in terms:
        x = _norm(str(x))
        if x and x in t:
            return True
    return False

def _pass_must_any_groups(text: str, must_any_groups: List[dict]) -> Tuple[bool, List[str]]:
    t = _norm(text)
    reasons = []
    for g in must_any_groups or []:
        any_terms = g.get("any") or []
        if any_terms and not _contains_any(t, [str(x) for x in any_terms]):
            reasons.append(f"missing_group:{g.get('name','group')}")
    return (len(reasons) == 0), reasons

def _cluster_match(cluster: str, name: str, subject_name: str, must_any_groups: List[dict]) -> Tuple[bool, List[str]]:
    text = f"{name} {subject_name}".strip()
    reasons = []
    if not _contains_any(text, CASE_INTENT_TERMS):
        return False, ["no_case_intent"]

    ok_groups, group_reasons = _pass_must_any_groups(text, must_any_groups)
    if not ok_groups and must_any_groups:
        reasons.extend(group_reasons)

    if cluster == "type":
        if not _contains_any(text, TPU_HINT_TERMS):
            reasons.append("no_tpu_signal")
        if not _contains_any(text, POCKET_HINT_TERMS):
            reasons.append("no_pocket_signal")

    if reasons:
        return False, reasons
    return True, []

def _ban_hit(text: str, ban_terms: List[str]) -> Optional[str]:
    t = _norm(text)
    for bt in ban_terms or []:
        bt_n = _norm(bt)
        if bt_n and bt_n in t:
            return bt_n
    return None

def _dedupe_by_imt_then_nm(items: List[dict]) -> List[dict]:
    seen_imt = set()
    seen_nm = set()
    out = []
    for it in items:
        imt = it.get("imt_id")
        nm = str(it.get("nm_id") or "")
        if imt is not None:
            if imt in seen_imt:
                continue
            seen_imt.add(imt)
        else:
            if nm in seen_nm:
                continue
            seen_nm.add(nm)
        out.append(it)
    return out


def stage_H_relevance_filter(
    out_dir: Path,
    *,
    resume: bool = False,
    strict: bool = False,
    min_keep_per_cluster: int = 12,
    max_keep_per_cluster: int = 28,
    max_per_seller: int = 3,
) -> Path:
    """
    Stage H (LOCAL):
    - rules-first KEEP/DROP for competitors per cluster
    - uses intent clusters' must_any_groups + ban_terms
    - if phone model missing -> relax model_tokens group only
    """
    manifest_path = out_dir / "run_manifest.json"
    intent_path = out_dir / "intent.jsonl"
    lite_path = out_dir / "competitor_lite.jsonl"
    if not manifest_path.exists():
        raise FileNotFoundError("run_manifest.json not found; run Stage A")
    if not intent_path.exists():
        raise FileNotFoundError("intent.jsonl not found; run Stage C")
    if not lite_path.exists():
        raise FileNotFoundError("competitor_lite.jsonl not found; run Stage G")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    # Load intent map (small scope)
    intent_map: Dict[str, dict] = {}
    with intent_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            j = json.loads(line)
            nm = str((j.get("meta") or {}).get("nm_id") or "")
            if nm:
                intent_map[nm] = j

    out_path = out_dir / "relevance.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    processed = 0

    with lite_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            rec = json.loads(line)
            meta = rec.get("meta") or {}
            own_nm = str(meta.get("nm_id") or "")
            if not own_nm or own_nm in done:
                continue

            intent_rec = intent_map.get(own_nm)
            if not intent_rec:
                if strict:
                    raise ValueError(f"Missing intent for nm_id={own_nm}")
                append_jsonl(out_path, {
                    "meta": {"schema_version": "2.0", "run_id": run_id, "nm_id": own_nm, "ts": _iso_utc(), "stage": "H"},
                    "clusters": [],
                    "notes": ["MISSING_INTENT_RECORD"],
                })
                processed += 1
                continue

            intent = intent_rec.get("intent") or {}
            clusters_cfg = {c.get("cluster"): c for c in (intent.get("clusters") or []) if isinstance(c, dict)}
            issues = list(intent.get("issues") or [])

            comp = rec.get("competitors") or {}
            merged = comp.get("merged") or []
            if not isinstance(merged, list):
                merged = []

            out_clusters = []
            notes = []
            if "MISSING_PHONE_MODEL" in issues:
                notes.append("LOW_MODEL_CONFIDENCE")

            for cluster in ("phone", "type"):
                cfg = clusters_cfg.get(cluster) or {}
                must_any_groups = list(cfg.get("must_any_groups") or [])
                ban_terms = list(cfg.get("ban_terms") or [])

                kept: List[dict] = []
                dropped: List[dict] = []
                seller_count: Dict[str, int] = {}

                for it in merged:
                    if not isinstance(it, dict):
                        continue
                    if cluster not in (it.get("clusters") or []):
                        continue

                    nm_id = str(it.get("nm_id") or "")
                    name = str(it.get("name") or "")
                    subj = str(it.get("subject_name") or "")
                    seller_key = str(it.get("seller_id") or it.get("seller_name") or "unknown")

                    if it.get("missing") is True:
                        dropped.append({"nm_id": nm_id, "reason": ["missing_card"]})
                        continue

                    text = f"{name} {subj}".strip()
                    bh = _ban_hit(text, ban_terms)
                    if bh:
                        dropped.append({"nm_id": nm_id, "reason": [f"ban_term:{bh}"]})
                        continue

                    # If phone model missing -> relax only model_tokens group
                    mg = must_any_groups
                    if "MISSING_PHONE_MODEL" in issues and mg:
                        mg = [g for g in mg if str(g.get("name") or "") != "model_tokens"]

                    ok, why = _cluster_match(cluster, name, subj, mg)
                    if not ok:
                        dropped.append({"nm_id": nm_id, "reason": why})
                        continue

                    # Diversify by seller
                    if seller_count.get(seller_key, 0) >= int(max_per_seller):
                        dropped.append({"nm_id": nm_id, "reason": ["seller_cap"]})
                        continue
                    seller_count[seller_key] = seller_count.get(seller_key, 0) + 1

                    kept.append(it)

                kept = _dedupe_by_imt_then_nm(kept)
                kept = kept[: max(0, int(max_keep_per_cluster))]

                if len(kept) < int(min_keep_per_cluster):
                    notes.append(f"LOW_CONF_{cluster}:kept={len(kept)}<min={min_keep_per_cluster}")

                kept_min = [{
                    "nm_id": str(x.get("nm_id") or ""),
                    "imt_id": x.get("imt_id"),
                    "seller_id": x.get("seller_id"),
                    "seller_name": x.get("seller_name"),
                    "name": x.get("name"),
                    "subject_name": x.get("subject_name"),
                    "rating": x.get("rating"),
                    "feedbacks": x.get("feedbacks"),
                    "price_rub": x.get("price_rub"),
                    "sale_price_rub": x.get("sale_price_rub"),
                    "total_quantity": x.get("total_quantity"),
                } for x in kept]

                out_clusters.append({
                    "cluster": cluster,
                    "kept": kept_min,
                    "dropped_sample": dropped[:50],
                    "stats": {
                        "in_cluster": sum(1 for x in merged if isinstance(x, dict) and cluster in (x.get("clusters") or [])),
                        "kept": len(kept_min),
                        "dropped": len(dropped),
                    },
                })

            # Unique imt_id pool to drive Stage I (review velocity)
            imt_ids = []
            for c in out_clusters:
                for x in c.get("kept") or []:
                    if x.get("imt_id") is not None:
                        imt_ids.append(int(x["imt_id"]))
            unique_imt = sorted(set(imt_ids))

            append_jsonl(out_path, {
                "meta": {
                    "schema_version": "2.0",
                    "run_id": run_id,
                    "nm_id": own_nm,
                    "vendor_code": str(meta.get("vendor_code") or ""),
                    "ts": _iso_utc(),
                    "stage": "H",
                },
                "clusters": out_clusters,
                "review_fetch_imt_ids": unique_imt,
                "notes": notes,
            })

            processed += 1

    return out_path

# =========================
# Stage I: MARKET PULSE (WB Necromancer v2)
# Reads:  run_manifest.json, relevance.jsonl
# Writes: market_pulse.jsonl (+ market_pulse_errors.jsonl)
# Caches: .wb_cache/reviews/imt_{imt_id}.json
# Network: YES (public feedbacks endpoints)
# =========================

# from __future__ import annotations

import json
import sys
import time
import random
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests


# --- helpers (keep consistent with previous stages) ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def write_json(path: Path, data: Any) -> None:
    ensure_dir(path.parent)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    tmp.replace(path)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def _utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(microsecond=0)


# --- HTTP ---
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
    "Content-Type": "application/json",
}

def _req_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    s.trust_env = True
    return s

def _backoff_sleep(attempt: int, base: float = 0.45, cap: float = 7.0) -> None:
    time.sleep(min(cap, base * (2 ** attempt) + random.random() * 0.25))

def post_json(
    sess: requests.Session,
    url: str,
    payload: dict,
    *,
    timeout: int,
    retries: int = 3
) -> Tuple[Optional[dict], Dict[str, Any]]:
    last_err = ""
    for a in range(retries + 1):
        try:
            r = sess.post(url, data=json.dumps(payload, ensure_ascii=False).encode("utf-8"), timeout=timeout)
            info = {"url": url, "status": r.status_code, "payload": payload}
            if r.status_code == 200:
                try:
                    return r.json(), info
                except Exception as e:
                    return None, {**info, "error": f"bad_json:{e!r}", "text_snip": r.text[:200]}
            if r.status_code in (429, 500, 502, 503, 504):
                _backoff_sleep(a)
                continue
            return None, {**info, "error": f"http_{r.status_code}", "text_snip": r.text[:200]}
        except Exception as e:
            last_err = repr(e)
            _backoff_sleep(a)
    return None, {"url": url, "status": None, "payload": payload, "error": last_err or "unknown"}


# --- date parsing (public feedback date formats are messy) ---
def _parse_dt(x: Any) -> Optional[datetime]:
    """
    English technical:
    Handles ISO strings like '2026-01-19T12:34:56Z' or with offset.
    Returns timezone-aware UTC datetime.
    """
    if x is None:
        return None
    if isinstance(x, (int, float)):
        # if unix ms/sec is ever used
        try:
            v = float(x)
            if v > 1e12:  # ms
                return datetime.fromtimestamp(v / 1000.0, tz=timezone.utc).replace(microsecond=0)
            if v > 1e9:  # sec
                return datetime.fromtimestamp(v, tz=timezone.utc).replace(microsecond=0)
        except Exception:
            return None

    s = str(x).strip()
    if not s:
        return None
    # normalize 'Z'
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).replace(microsecond=0)
    except Exception:
        return None


def _extract_feedback_list(js: Any) -> List[dict]:
    """
    English technical:
    Supports responses like:
      - { feedbacks: [...] }
      - { data: { feedbacks: [...] } }
      - [...] (rare)
    """
    if js is None:
        return []
    if isinstance(js, list):
        return [x for x in js if isinstance(x, dict)]
    if isinstance(js, dict):
        if isinstance(js.get("feedbacks"), list):
            return [x for x in js["feedbacks"] if isinstance(x, dict)]
        data = js.get("data")
        if isinstance(data, dict) and isinstance(data.get("feedbacks"), list):
            return [x for x in data["feedbacks"] if isinstance(x, dict)]
    return []


@dataclass
class ReviewSummary:
    imt_id: int
    recent_30: int
    recent_90: int
    days_since_last: Optional[int]
    newest_dt: Optional[str]
    oldest_dt: Optional[str]
    pages_fetched: int
    reviews_seen: int
    source_url: str
    warnings: List[str]


def fetch_review_summary_for_imt(
    sess: requests.Session,
    imt_id: int,
    *,
    base_urls: List[str],
    timeout: int,
    take: int = 30,
    max_pages: int = 40,
    max_skip: int = 1000,  # avoid known 400s when skip too large
    early_stop_days: int = 90,
    sleep_s: float = 0.25,
) -> Tuple[Optional[ReviewSummary], Dict[str, Any]]:
    """
    English technical:
    Fetches public feedback pages ordered by dateDesc.
    Early-stop when the oldest item in a page is older than early_stop_days.
    """
    now = _utc_now()
    cut_30 = now - timedelta(days=30)
    cut_90 = now - timedelta(days=90)

    warnings: List[str] = []
    all_dts: List[datetime] = []
    recent_30 = 0
    recent_90 = 0
    pages_fetched = 0
    reviews_seen = 0

    used_url = ""
    last_info: Dict[str, Any] = {}

    # try each base URL until one works
    for base in base_urls:
        ok_any = False
        pages_fetched = 0
        reviews_seen = 0
        all_dts = []
        recent_30 = 0
        recent_90 = 0
        used_url = base
        last_info = {"imt_id": imt_id, "base_url": base, "pages": []}

        for page in range(max_pages):
            skip = page * take
            if skip > max_skip:
                warnings.append("skip_cap_reached")
                break

            payload = {"imtId": int(imt_id), "take": int(take), "skip": int(skip), "order": "dateDesc"}
            js, info = post_json(sess, base, payload, timeout=timeout, retries=3)
            last_info["pages"].append(info)

            if js is None:
                # if the first page fails, try next base URL
                if page == 0:
                    break
                warnings.append("page_fetch_failed")
                break

            items = _extract_feedback_list(js)
            pages_fetched += 1
            if not items:
                ok_any = True
                break

            page_dts: List[datetime] = []
            for it in items:
                dt = _parse_dt(it.get("createdDate") or it.get("createdDateTime") or it.get("date") or it.get("created"))
                if dt is None:
                    continue
                page_dts.append(dt)
                all_dts.append(dt)
                reviews_seen += 1
                if dt >= cut_30:
                    recent_30 += 1
                if dt >= cut_90:
                    recent_90 += 1

            ok_any = True

            # early stop check on page oldest
            if page_dts:
                oldest = min(page_dts)
                if oldest < cut_90:
                    break

            time.sleep(max(0.0, float(sleep_s)))

        if ok_any:
            # we got at least one successful page from this base
            break

    if not all_dts:
        # no reviews or endpoint failed
        # distinguish: if first page failed everywhere => endpoint issue
        # We can infer from last_info if it has pages with errors.
        warn2 = warnings[:]
        if not used_url:
            warn2.append("no_base_url_worked")
        return ReviewSummary(
            imt_id=int(imt_id),
            recent_30=0,
            recent_90=0,
            days_since_last=None,
            newest_dt=None,
            oldest_dt=None,
            pages_fetched=pages_fetched,
            reviews_seen=reviews_seen,
            source_url=used_url,
            warnings=warn2 or ["no_reviews_or_unreachable"],
        ), last_info

    newest = max(all_dts)
    oldest = min(all_dts)
    days_since_last = int((_utc_now() - newest).total_seconds() // 86400)

    return ReviewSummary(
        imt_id=int(imt_id),
        recent_30=int(recent_30),
        recent_90=int(recent_90),
        days_since_last=days_since_last,
        newest_dt=newest.isoformat(),
        oldest_dt=oldest.isoformat(),
        pages_fetched=int(pages_fetched),
        reviews_seen=int(reviews_seen),
        source_url=used_url,
        warnings=warnings,
    ), last_info


def _median_int(vals: List[int]) -> Optional[float]:
    if not vals:
        return None
    vs = sorted(vals)
    n = len(vs)
    mid = n // 2
    if n % 2 == 1:
        return float(vs[mid])
    return (vs[mid - 1] + vs[mid]) / 2.0


def stage_I_market_pulse(
    out_dir: Path,
    *,
    timeout: int = 30,
    sleep_s: float = 0.25,
    resume: bool = False,
    strict: bool = False,
    cache_ttl_hours: int = 24,
    take: int = 30,
    max_pages: int = 40,
    max_skip: int = 1000,
    early_stop_days: int = 90,
) -> Path:
    """
    Stage I:
    - Reads relevance.jsonl (Stage H) for per-SKU competitor imt_ids
    - Fetches public feedback summaries per imt_id with caching
    - Outputs per-SKU per-cluster aggregated Market Pulse metrics
    """
    manifest_path = out_dir / "run_manifest.json"
    rel_path = out_dir / "relevance.jsonl"
    if not manifest_path.exists():
        raise FileNotFoundError("run_manifest.json not found; run Stage A")
    if not rel_path.exists():
        raise FileNotFoundError("relevance.jsonl not found; run Stage H")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    cfg = (manifest.get("config") or {})
    reviews_cfg = cfg.get("reviews") or {}

    # Default public endpoints list; keep configurable because WB likes breaking stuff.
    base_urls = list(reviews_cfg.get("base_urls") or [
        "https://public-feedbacks.wildberries.ru/api/v1/feedbacks/site",
        "https://public-feedbacks.wildberries.ru/api/v1/feedbacks",
    ])

    out_path = out_dir / "market_pulse.jsonl"
    err_path = out_dir / "market_pulse_errors.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    sess = _req_session()

    cache_dir = out_dir / ".wb_cache" / "reviews"
    ensure_dir(cache_dir)

    def cache_path_for(imt_id: int) -> Path:
        return cache_dir / f"imt_{int(imt_id)}.json"

    def cache_is_fresh(p: Path) -> bool:
        if not p.exists():
            return False
        try:
            js = read_json(p)
            ts = js.get("fetched_at")
            dt = _parse_dt(ts)
            if not dt:
                return False
            age_h = (_utc_now() - dt).total_seconds() / 3600.0
            return age_h <= float(cache_ttl_hours)
        except Exception:
            return False

    def load_cached_summary(p: Path) -> Optional[ReviewSummary]:
        try:
            js = read_json(p)
            s = js.get("summary") or {}
            return ReviewSummary(
                imt_id=int(s["imt_id"]),
                recent_30=int(s.get("recent_30") or 0),
                recent_90=int(s.get("recent_90") or 0),
                days_since_last=s.get("days_since_last"),
                newest_dt=s.get("newest_dt"),
                oldest_dt=s.get("oldest_dt"),
                pages_fetched=int(s.get("pages_fetched") or 0),
                reviews_seen=int(s.get("reviews_seen") or 0),
                source_url=str(s.get("source_url") or ""),
                warnings=list(s.get("warnings") or []),
            )
        except Exception:
            return None

    processed = 0

    with rel_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            rec = json.loads(line)
            meta = rec.get("meta") or {}
            own_nm = str(meta.get("nm_id") or "")
            if not own_nm or own_nm in done:
                continue

            clusters = rec.get("clusters") or []
            # Build cluster -> imt_ids from kept competitors
            cluster_imts: Dict[str, List[int]] = {"phone": [], "type": []}
            cluster_kept_nm: Dict[str, List[str]] = {"phone": [], "type": []}

            for c in clusters:
                cl = str(c.get("cluster") or "")
                if cl not in ("phone", "type"):
                    continue
                kept = c.get("kept") or []
                for it in kept:
                    imt = it.get("imt_id")
                    nm = str(it.get("nm_id") or "")
                    if nm and nm.isdigit():
                        cluster_kept_nm[cl].append(nm)
                    try:
                        if imt is not None:
                            cluster_imts[cl].append(int(imt))
                    except Exception:
                        pass

            # Dedupe keep order
            for cl in ("phone", "type"):
                seen = set()
                out = []
                for x in cluster_imts[cl]:
                    if x in seen:
                        continue
                    seen.add(x)
                    out.append(x)
                cluster_imts[cl] = out
                cluster_kept_nm[cl] = list(dict.fromkeys(cluster_kept_nm[cl]))

            # Union imt_ids for fetching
            all_imts: List[int] = []
            for cl in ("phone", "type"):
                all_imts.extend(cluster_imts[cl])
            all_imts = list(dict.fromkeys(all_imts))

            summaries: Dict[int, ReviewSummary] = {}
            fetch_notes: List[str] = []

            for imt_id in all_imts:
                cp = cache_path_for(imt_id)

                try:
                    if cache_is_fresh(cp):
                        s = load_cached_summary(cp)
                        if s:
                            summaries[imt_id] = s
                            continue

                    s, debug = fetch_review_summary_for_imt(
                        sess,
                        imt_id,
                        base_urls=base_urls,
                        timeout=timeout,
                        take=take,
                        max_pages=max_pages,
                        max_skip=max_skip,
                        early_stop_days=early_stop_days,
                        sleep_s=sleep_s,
                    )
                    if s is None:
                        raise RuntimeError("review_summary_none")

                    summaries[imt_id] = s

                    # Save cache (summary + minimal debug)
                    write_json(cp, {
                        "imt_id": int(imt_id),
                        "fetched_at": _iso_utc(),
                        "summary": {
                            "imt_id": s.imt_id,
                            "recent_30": s.recent_30,
                            "recent_90": s.recent_90,
                            "days_since_last": s.days_since_last,
                            "newest_dt": s.newest_dt,
                            "oldest_dt": s.oldest_dt,
                            "pages_fetched": s.pages_fetched,
                            "reviews_seen": s.reviews_seen,
                            "source_url": s.source_url,
                            "warnings": s.warnings,
                        },
                        "debug": debug,
                    })

                except Exception as e:
                    append_jsonl(err_path, {
                        "meta": {"schema_version": "2.0", "run_id": run_id, "nm_id": own_nm, "ts": _iso_utc(), "stage": "I"},
                        "imt_id": int(imt_id),
                        "error": repr(e),
                    })
                    fetch_notes.append(f"imt_fetch_error:{imt_id}")
                    if strict:
                        raise

            # Aggregate per cluster
            out_clusters: List[dict] = []
            for cl in ("phone", "type"):
                imts = cluster_imts[cl]
                rows = [summaries.get(i) for i in imts if summaries.get(i) is not None]

                recent30 = [r.recent_30 for r in rows]
                recent90 = [r.recent_90 for r in rows]
                dsl = [r.days_since_last for r in rows if r.days_since_last is not None]

                out_clusters.append({
                    "cluster": cl,
                    "imt_n": len(imts),
                    "imt_with_data_n": len(rows),
                    "recent_30_sum": int(sum(recent30)) if recent30 else 0,
                    "recent_90_sum": int(sum(recent90)) if recent90 else 0,
                    "recent_30_median": _median_int(recent30),
                    "recent_90_median": _median_int(recent90),
                    "days_since_last_median": _median_int([int(x) for x in dsl]) if dsl else None,
                    "days_since_last_min": int(min(dsl)) if dsl else None,
                    "days_since_last_max": int(max(dsl)) if dsl else None,
                    "sample": [
                        {
                            "imt_id": r.imt_id,
                            "recent_30": r.recent_30,
                            "recent_90": r.recent_90,
                            "days_since_last": r.days_since_last,
                            "newest_dt": r.newest_dt,
                            "warnings": r.warnings[:3],
                        }
                        for r in rows[: min(12, len(rows))]
                    ],
                })

            append_jsonl(out_path, {
                "meta": {
                    "schema_version": "2.0",
                    "run_id": run_id,
                    "nm_id": own_nm,
                    "vendor_code": str(meta.get("vendor_code") or ""),
                    "ts": _iso_utc(),
                    "stage": "I",
                },
                "market_pulse": {
                    "clusters": out_clusters,
                    "source": {"base_urls": base_urls, "take": int(take), "max_pages": int(max_pages), "early_stop_days": int(early_stop_days)},
                },
                "notes": fetch_notes,
            })

            processed += 1
            if processed % 10 == 0:
                print(f"[I] processed: {processed}", file=sys.stderr)

    print(f"[I] processed: {processed}")
    return out_path

# =========================
# Stage registry / UX (wb_revive.py style)
# =========================

STAGE_ORDER = ["A","B","C","D","E","F","G","H","I","J","K","L","M"]

STAGE_META = {
    "A": {"title": "INPUT + Manifest", "network": "LOCAL", "llm_flag": None},
    "B": {"title": "OWN FETCH (WB)", "network": "WB", "llm_flag": None},
    "C": {"title": "INTENT EXTRACT (rules-first)", "network": "LOCAL", "llm_flag": None},
    "D": {"title": "QUERY BUILD (2 clusters)", "network": "LOCAL", "llm_flag": None},  # LLM hook later
    "E": {"title": "SERP SNAPSHOT + VALIDATION (WB)", "network": "WB", "llm_flag": None},
    "F": {"title": "COMPETITOR POOL (LOCAL)", "network": "LOCAL", "llm_flag": None},
    "G": {"title": "LITE FETCH (WB)", "network": "WB", "llm_flag": None},
    "H": {"title": "RELEVANCE FILTER (rules-first + optional LLM)", "network": "LOCAL", "llm_flag": "use_llm_h"},
    "I": {"title": "MARKET PULSE (WB reviews)", "network": "WB", "llm_flag": None},
    "J": {"title": "SUPPLY/STRUCTURE (LOCAL facts)", "network": "LOCAL", "llm_flag": None},
    "K": {"title": "CLUSTER VERDICTS (rules)", "network": "LOCAL", "llm_flag": None},
    "L": {"title": "FINAL DECISION (optional LLM phrasing)", "network": "LLM", "llm_flag": "use_llm_l"},
    "M": {"title": "REPORTS (XLSX + HTML + optional exec summary)", "network": "LOCAL", "llm_flag": "use_llm_m"},
}

def _stage_title(code: str) -> str:
    return STAGE_META.get(code, {}).get("title") or code

def _stage_network(code: str) -> str:
    return STAGE_META.get(code, {}).get("network") or "LOCAL"

def _stage_llm_enabled(code: str, args) -> bool:
    flag = STAGE_META.get(code, {}).get("llm_flag")
    return bool(flag and getattr(args, flag, False))

def _vpn_hint(code: str, args) -> str:
    net = _stage_network(code)
    if net == "LLM" and _stage_llm_enabled(code, args):
        return "возможно нужен (LLM API)"
    if net == "WB":
        return "не нужен (WB)"
    return "не нужен"

def print_stage_banner(code: str, args, *, out_dir: str = "") -> None:
    net = _stage_network(code)
    llm = "ON" if _stage_llm_enabled(code, args) else "OFF"
    title = _stage_title(code)
    hint = _vpn_hint(code, args)
    line = "=" * 86
    print(line)
    print(f"[Stage {code}] {title}")
    print(f"  network: {net} | LLM: {llm} | VPN hint: {hint}")
    if out_dir:
        print(f"  out_dir: {out_dir}")
    print(line)

# =========================
# Stage L: FINAL DECISION (WB Necromancer v2)
# Reads:  run_manifest.json, intent.jsonl, cluster_verdicts.jsonl
# Writes: decisions.jsonl
# Network: LOCAL (optional LLM for phrasing)
# =========================

# from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# --- minimal helpers ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def _load_jsonl_map(path: Path, key_field: str = "nm_id") -> Dict[str, dict]:
    m: Dict[str, dict] = {}
    if not path.exists():
        return m
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
            except Exception:
                continue
            nm = str((j.get("meta") or {}).get(key_field) or "")
            if nm:
                m[nm] = j
    return m


def _get_cluster_row(verdict_rec: dict, cluster: str) -> dict:
    rows = verdict_rec.get("cluster_verdicts") or []
    for r in rows:
        if str(r.get("cluster") or "") == cluster:
            return r
    return {}

def _get_karma(intent_rec: dict) -> Tuple[Optional[float], Optional[int]]:
    k = (intent_rec.get("intent") or {}).get("karma") or {}
    rating = k.get("rating")
    feedbacks = k.get("feedbacks")
    try:
        rating = float(rating) if rating is not None else None
    except Exception:
        rating = None
    try:
        feedbacks = int(float(feedbacks)) if feedbacks is not None else None
    except Exception:
        feedbacks = None
    return rating, feedbacks

def _karma_is_toxic(
    rating: Optional[float],
    feedbacks: Optional[int],
    *,
    min_fb: int,
    min_rating: float
) -> bool:
    # “токсична” = низкий рейтинг при достаточном числе отзывов
    if rating is None or feedbacks is None:
        return False
    return feedbacks >= int(min_fb) and rating < float(min_rating)

def _backlog_from_issues(issues: List[str], *, decision: str) -> List[str]:
    tasks: List[str] = []
    issues = [str(x) for x in (issues or [])]

    if "MISSING_PHONE_MODEL" in issues:
        tasks.append("Уточнить и зафиксировать модель телефона (в названии, характеристиках, описании).")
    if "NO_TPU_SIGNAL" in issues:
        tasks.append("Явно указать материал TPU и синонимы (силикон, термополиуретан) в тексте и атрибутах.")
    if "NO_POCKET_SIGNAL" in issues:
        tasks.append("Явно указать карман или картхолдер (карман для карт, слот под карту) в тексте и на фото.")

    if decision in ("REVIVE_REWORK", "CLONE_NEW_CARD"):
        tasks.extend([
            "Пересобрать SEO: заголовок до 60 символов, ключи без спама, заполнить характеристики под фильтры.",
            "Проверить фото: добавить кадр с карманом для карты и материалом, привести к единому стилю.",
        ])

    if decision == "CLONE_NEW_CARD":
        tasks.append("Создать новую карточку и перенести лучший контент без токсичной истории отзывов.")

    return tasks[:12]


def stage_L_final_decision(
    out_dir: Path,
    *,
    resume: bool = False,
    use_llm_l: bool = False,
    llm_provider: str = "openrouter",
    llm_model: str = "openai/gpt-4o-mini",
    api_key: str = "",
    llm_base_url: str = "",
    llm_timeout: int = 60,
    llm_max_tokens: int = 1200,
    llm_temperature: float = 0.2,
) -> Path:
    """
    Stage L:
    - Applies v2 decision matrix (rules-first)
    - Optional: LLM phrasing based on FACTS (must not change decision)
    Output enum: DROP | REVIVE_FAST | REVIVE_REWORK | CLONE_NEW_CARD
    """
    manifest_path = out_dir / "run_manifest.json"
    intent_path = out_dir / "intent.jsonl"
    cv_path = out_dir / "cluster_verdicts.jsonl"

    if not manifest_path.exists():
        raise FileNotFoundError("run_manifest.json not found; run Stage A")
    if not intent_path.exists():
        raise FileNotFoundError("intent.jsonl not found; run Stage C")
    if not cv_path.exists():
        raise FileNotFoundError("cluster_verdicts.jsonl not found; run Stage K")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    # Configurable karma thresholds (defaults match v2 intent)
    cfg = manifest.get("config") or {}
    karma_cfg = cfg.get("karma") or {}
    karma_min_fb = int(karma_cfg.get("min_feedbacks") or 25)
    karma_toxic_rating_lt = float(karma_cfg.get("toxic_rating_lt") or 4.4)

    intent_map = _load_jsonl_map(intent_path, "nm_id")
    cv_map = _load_jsonl_map(cv_path, "nm_id")

    out_path = out_dir / "decisions.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    processed = 0

    for own_nm, cv in cv_map.items():
        if not own_nm or own_nm in done:
            continue

        intent_rec = intent_map.get(own_nm) or {}
        vendor_code = str(
            ((cv.get("meta") or {}).get("vendor_code")) or
            ((intent_rec.get("meta") or {}).get("vendor_code")) or
            ""
        )

        phone_row = _get_cluster_row(cv, "phone")
        type_row = _get_cluster_row(cv, "type")

        phone_status = str(phone_row.get("status") or "DEAD")
        type_status = str(type_row.get("status") or "DEAD")

        rating, feedbacks = _get_karma(intent_rec)
        issues = list(((intent_rec.get("intent") or {}).get("issues")) or [])
        extracted = ((intent_rec.get("intent") or {}).get("extracted")) or {}

        toxic = _karma_is_toxic(
            rating, feedbacks,
            min_fb=karma_min_fb,
            min_rating=karma_toxic_rating_lt
        )

        risk_flags = sorted(set(
            (cv.get("risk_flags") or []) +
            (phone_row.get("risk_flags") or []) +
            (type_row.get("risk_flags") or [])
        ))
        notes = list(cv.get("notes") or [])

        decision = "DROP"
        alt_strategy_ru: Optional[str] = None

        # =========================
        # Decision matrix v2
        # =========================
        if phone_status == "DEAD":
            decision = "DROP"
            risk_flags = sorted(set(risk_flags + ["PHONE_MARKET_DEAD"]))
        else:
            # phone is ALIVE or SLOW
            if type_status == "DEAD":
                decision = "DROP"
                risk_flags = sorted(set(risk_flags + ["TYPE_MARKET_DEAD"]))
                alt_strategy_ru = (
                    "Модель телефона выглядит живой, но тип «TPU + карман для карты» не подтверждён рынком. "
                    "Если хочешь работать по модели, рассмотри другие типы чехлов (обычный TPU, противоударный, книжка)."
                )
            else:
                # type is ALIVE or SLOW
                if toxic:
                    decision = "CLONE_NEW_CARD"
                    risk_flags = sorted(set(risk_flags + ["TOXIC_KARMA"]))
                else:
                    content_debt = False

                    if "MISSING_PHONE_MODEL" in issues:
                        content_debt = True
                    if "NO_TPU_SIGNAL" in issues or "NO_POCKET_SIGNAL" in issues:
                        content_debt = True

                    # risk flags that usually force “думать, а не жать кнопку”
                    if "LOW_CONFIDENCE" in risk_flags:
                        content_debt = True
                    if "DUMPING_PRESSURE" in risk_flags or "MONOPOLY_DANGER" in risk_flags:
                        content_debt = True

                    if type_status == "SLOW":
                        content_debt = True
                        risk_flags = sorted(set(risk_flags + ["SLOW_DEMAND"]))

                    decision = "REVIVE_REWORK" if content_debt else "REVIVE_FAST"

        backlog = _backlog_from_issues(issues, decision=decision)

        # =========================
        # Deterministic rationale
        # =========================
        phone_conf = str(phone_row.get("confidence") or "")
        type_conf = str(type_row.get("confidence") or "")

        karma_part = ""
        if rating is not None and feedbacks is not None:
            karma_part = f"Карма карточки: рейтинг {rating:.2f}, отзывов {feedbacks}."
        elif rating is not None:
            karma_part = f"Карма карточки: рейтинг {rating:.2f}."
        elif feedbacks is not None:
            karma_part = f"Карма карточки: отзывов {feedbacks}."

        rationale_ru = (
            f"Рынок модели (phone) = {phone_status} (conf={phone_conf}); "
            f"рынок типа TPU+карман (type) = {type_status} (conf={type_conf}). "
            f"{karma_part} "
        ).strip()

        if decision == "DROP":
            if phone_status == "DEAD":
                rationale_ru += " Общий рынок чехлов на модель не выглядит живым: оживление SKU не окупится."
            else:
                rationale_ru += " Модель в целом жива, но спрос на «TPU + карман» не подтверждён: обычно это повод не тратить ресурсы."
        elif decision == "CLONE_NEW_CARD":
            rationale_ru += " Типовой рынок подтверждён, но история карточки токсична: клонирование в новую карточку рациональнее."
        elif decision == "REVIVE_FAST":
            rationale_ru += " Оба рынка выглядят живыми и без крупных рисков: можно оживлять быстро."
        else:
            rationale_ru += " Рынки не мёртвые, но есть риски или контент-долг: нужна переработка перед оживлением."

        # =========================
        # Optional LLM phrasing (FACTS-only)
        # =========================
        llm_used = False
        llm_debug: Dict[str, Any] = {}

        if use_llm_l:
            # Expect call_llm_json() to exist in the final script (borrow from wb_revive.py).
            if "call_llm_json" not in globals():
                raise RuntimeError("use_llm_l=True but call_llm_json() is not defined in this script.")

            facts = {
                "decision_fixed": decision,
                "phone_market": phone_row,
                "type_market": type_row,
                "karma": {"rating": rating, "feedbacks": feedbacks, "toxic": toxic},
                "risk_flags": risk_flags,
                "issues": issues,
                "extracted": extracted,
                "alt_strategy_ru": alt_strategy_ru,
            }

            messages = [
                {"role": "system", "content": (
                    "You are a strict analyst. Output JSON only. "
                    "Do not invent numbers. Use only FACTS. "
                    "You must not change decision_fixed."
                )},
                {"role": "user", "content": (
                    "FACTS (JSON):\n"
                    f"{json.dumps(facts, ensure_ascii=False)}\n\n"
                    "Return JSON object with keys: "
                    "rationale_ru (string), backlog_ru (array of strings), risk_flags (array of EN UPPER_SNAKE_CASE), "
                    "alt_strategy_ru (string or null). Keep it concise."
                )},
            ]

            obj, dbg = call_llm_json(
                provider=llm_provider,
                model=llm_model,
                api_key=api_key,
                messages=messages,
                base_url=llm_base_url,
                timeout_sec=llm_timeout,
                max_tokens=llm_max_tokens,
                temperature=llm_temperature,
                force_json=True,
            )
            llm_used = True
            llm_debug = dbg or {}

            if isinstance(obj, dict):
                rr = obj.get("rationale_ru")
                if isinstance(rr, str) and rr.strip():
                    rationale_ru = rr.strip()

                bl = obj.get("backlog_ru")
                if isinstance(bl, list) and bl:
                    backlog = [str(x) for x in bl if str(x).strip()][:12]

                rf = obj.get("risk_flags")
                if isinstance(rf, list) and rf:
                    risk_flags = sorted(set(risk_flags + [str(x) for x in rf if str(x).strip()]))

                ar = obj.get("alt_strategy_ru")
                if isinstance(ar, str) and ar.strip():
                    alt_strategy_ru = ar.strip()

        append_jsonl(out_path, {
            "meta": {
                "schema_version": "2.0",
                "run_id": run_id,
                "nm_id": str(own_nm),
                "vendor_code": vendor_code,
                "ts": _iso_utc(),
                "stage": "L",
            },
            "decision": {
                "final": decision,  # DROP|REVIVE_FAST|REVIVE_REWORK|CLONE_NEW_CARD
                "phone_status": phone_status,
                "type_status": type_status,
                "confidence": {
                    "phone": phone_row.get("confidence"),
                    "type": type_row.get("confidence"),
                },
                "karma": {
                    "rating": rating,
                    "feedbacks": feedbacks,
                    "toxic": toxic,
                    "toxic_rule": {"min_feedbacks": karma_min_fb, "rating_lt": karma_toxic_rating_lt},
                },
                "risk_flags": risk_flags,
                "alt_strategy_ru": alt_strategy_ru,
            },
            "rationale_ru": rationale_ru,
            "backlog_ru": backlog,
            "evidence": {
                "cluster_verdicts": cv.get("cluster_verdicts"),
                "facts_phone": phone_row.get("facts"),
                "facts_type": type_row.get("facts"),
                "issues": issues,
                "notes": notes,
            },
            "llm": {
                "used": bool(llm_used),
                "provider": llm_provider if llm_used else "",
                "model": llm_model if llm_used else "",
                "debug": llm_debug,
            }
        })

        processed += 1

    return out_path


### далее может быть перепутан порядок, проверить внимательно!!!!
# =========================
# Stage registry / UX (wb_revive.py style)
# =========================

STAGE_ORDER = ["A","B","C","D","E","F","G","H","I","J","K","L","M"]

STAGE_META = {
    "A": {"title": "INPUT + Manifest", "network": "LOCAL", "llm_flag": None},
    "B": {"title": "OWN FETCH (WB)", "network": "WB", "llm_flag": None},
    "C": {"title": "INTENT EXTRACT (rules-first)", "network": "LOCAL", "llm_flag": None},
    "D": {"title": "QUERY BUILD (2 clusters)", "network": "LOCAL", "llm_flag": None},  # LLM hook later
    "E": {"title": "SERP SNAPSHOT + VALIDATION (WB)", "network": "WB", "llm_flag": None},
    "F": {"title": "COMPETITOR POOL (LOCAL)", "network": "LOCAL", "llm_flag": None},
    "G": {"title": "LITE FETCH (WB)", "network": "WB", "llm_flag": None},
    "H": {"title": "RELEVANCE FILTER (rules-first + optional LLM)", "network": "LOCAL", "llm_flag": "use_llm_h"},
    "I": {"title": "MARKET PULSE (WB reviews)", "network": "WB", "llm_flag": None},
    "J": {"title": "SUPPLY/STRUCTURE (LOCAL facts)", "network": "LOCAL", "llm_flag": None},
    "K": {"title": "CLUSTER VERDICTS (rules)", "network": "LOCAL", "llm_flag": None},
    "L": {"title": "FINAL DECISION (optional LLM phrasing)", "network": "LLM", "llm_flag": "use_llm_l"},
    "M": {"title": "REPORTS (XLSX + HTML + optional exec summary)", "network": "LOCAL", "llm_flag": "use_llm_m"},
}

def _stage_title(code: str) -> str:
    return STAGE_META.get(code, {}).get("title") or code

def _stage_network(code: str) -> str:
    return STAGE_META.get(code, {}).get("network") or "LOCAL"

def _stage_llm_enabled(code: str, args) -> bool:
    flag = STAGE_META.get(code, {}).get("llm_flag")
    return bool(flag and getattr(args, flag, False))

def _vpn_hint(code: str, args) -> str:
    net = _stage_network(code)
    if net == "LLM" and _stage_llm_enabled(code, args):
        return "возможно нужен (LLM API)"
    if net == "WB":
        return "не нужен (WB)"
    return "не нужен"

def print_stage_banner(code: str, args, *, out_dir: str = "") -> None:
    net = _stage_network(code)
    llm = "ON" if _stage_llm_enabled(code, args) else "OFF"
    title = _stage_title(code)
    hint = _vpn_hint(code, args)
    line = "=" * 86
    print(line)
    print(f"[Stage {code}] {title}")
    print(f"  network: {net} | LLM: {llm} | VPN hint: {hint}")
    if out_dir:
        print(f"  out_dir: {out_dir}")
    print(line)

# =========================
# Stage J: SUPPLY/STRUCTURE (WB Necromancer v2)
# Reads:  run_manifest.json, relevance.jsonl
# Writes: supply_structure.jsonl
# Network: NO (LOCAL)
# =========================

# from __future__ import annotations

import json
import math
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# --- minimal helpers ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def _safe_float(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        v = float(x)
        if not math.isfinite(v):
            return None
        return v
    except Exception:
        return None

def _safe_int(x: Any) -> Optional[int]:
    try:
        if x is None:
            return None
        v = int(float(x))
        return v
    except Exception:
        return None


# --- robust stats ---
def _quantile(sorted_vals: List[float], q: float) -> Optional[float]:
    if not sorted_vals:
        return None
    n = len(sorted_vals)
    if n == 1:
        return float(sorted_vals[0])
    q = max(0.0, min(1.0, float(q)))
    pos = q * (n - 1)
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return float(sorted_vals[lo])
    w = pos - lo
    return float(sorted_vals[lo] * (1 - w) + sorted_vals[hi] * w)

def _median(vals: List[float]) -> Optional[float]:
    if not vals:
        return None
    s = sorted(vals)
    return _quantile(s, 0.5)

def _trim(vals: List[float], trim: float = 0.10) -> List[float]:
    if not vals:
        return []
    s = sorted(vals)
    n = len(s)
    k = int(math.floor(n * float(trim)))
    if 2 * k >= n:
        return s
    return s[k:n-k]

def _mad(vals: List[float], med: float) -> Optional[float]:
    if not vals:
        return None
    dev = [abs(v - med) for v in vals]
    return _median(dev)

def _outlier_counts(vals: List[float], med: float) -> Tuple[int, int, float]:
    """
    English technical:
    Robust z-score based on MAD: z = 0.6745*(x-med)/MAD.
    """
    if not vals:
        return 0, 0, 0.0
    mad = _mad(vals, med)
    if mad is None or mad == 0:
        return 0, 0, 0.0
    low = 0
    high = 0
    for v in vals:
        z = 0.6745 * (v - med) / mad
        if z < -3.5:
            low += 1
        elif z > 3.5:
            high += 1
    rate = (low + high) / max(1, len(vals))
    return low, high, round(rate, 3)


# --- concentration metrics ---
def _hhi_from_counts(counts: List[int]) -> Optional[float]:
    if not counts:
        return None
    tot = sum(counts)
    if tot <= 0:
        return None
    hhi = 0.0
    for c in counts:
        s = c / tot
        hhi += s * s
    return round(hhi, 4)

def _top1_share_from_counts(counts: List[int]) -> Optional[float]:
    if not counts:
        return None
    tot = sum(counts)
    if tot <= 0:
        return None
    return round(max(counts) / tot, 4)


def stage_J_supply_structure(
    out_dir: Path,
    *,
    resume: bool = False,
    min_n: int = 8,
) -> Path:
    """
    Stage J:
    - price robust stats per cluster (median, p10/p90, trimmed median, outliers)
    - stock proxy (qty median + qty concentration if available)
    - seller concentration (unique sellers, top1 share, HHI)
    - risk flags: DUMPING_PRESSURE, MONOPOLY_DANGER, LOW_CONFIDENCE
    """
    manifest_path = out_dir / "run_manifest.json"
    rel_path = out_dir / "relevance.jsonl"
    if not manifest_path.exists():
        raise FileNotFoundError("run_manifest.json not found; run Stage A")
    if not rel_path.exists():
        raise FileNotFoundError("relevance.jsonl not found; run Stage H")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    out_path = out_dir / "supply_structure.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    processed = 0

    with rel_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            rec = json.loads(line)
            meta = rec.get("meta") or {}
            own_nm = str(meta.get("nm_id") or "")
            if not own_nm or own_nm in done:
                continue

            clusters = rec.get("clusters") or []
            notes_in = list(rec.get("notes") or [])

            out_clusters: List[dict] = []
            global_flags: List[str] = []

            for cl in ("phone", "type"):
                c = next((x for x in clusters if str(x.get("cluster") or "") == cl), None)
                kept = (c or {}).get("kept") or []

                # --- PRICE ---
                prices: List[float] = []
                for it in kept:
                    sp = _safe_float(it.get("sale_price_rub"))
                    pr = _safe_float(it.get("price_rub"))
                    use = sp if sp is not None else pr
                    if use is not None and use > 0:
                        prices.append(use)

                prices_s = sorted(prices)
                p_med = _quantile(prices_s, 0.5) if prices_s else None
                p10 = _quantile(prices_s, 0.10) if prices_s else None
                p90 = _quantile(prices_s, 0.90) if prices_s else None
                p_trim_med = None
                if prices_s:
                    t = _trim(prices_s, 0.10)
                    p_trim_med = _median(t) if t else p_med

                low_out = high_out = 0
                out_rate = 0.0
                if p_med is not None and prices_s:
                    low_out, high_out, out_rate = _outlier_counts(prices_s, p_med)

                # --- SELLERS (counts) ---
                sellers = []
                for it in kept:
                    sid = it.get("seller_id")
                    sn = str(it.get("seller_name") or "")
                    key = f"id:{sid}" if sid is not None else f"name:{sn.strip().lower()}" if sn else "unknown"
                    sellers.append(key)

                seller_counts: Dict[str, int] = {}
                for s in sellers:
                    seller_counts[s] = seller_counts.get(s, 0) + 1

                unique_sellers = len([k for k in seller_counts.keys() if k != "unknown"])
                cnts = list(seller_counts.values())
                top1_share = _top1_share_from_counts(cnts)
                hhi = _hhi_from_counts(cnts)

                # --- STOCK PROXY ---
                qtys: List[int] = []
                qty_by_seller: Dict[str, int] = {}
                for it in kept:
                    q = _safe_int(it.get("total_quantity"))
                    if q is None or q < 0:
                        continue
                    qtys.append(q)
                    sid = it.get("seller_id")
                    sn = str(it.get("seller_name") or "")
                    key = f"id:{sid}" if sid is not None else f"name:{sn.strip().lower()}" if sn else "unknown"
                    qty_by_seller[key] = qty_by_seller.get(key, 0) + q

                qtys_s = sorted(qtys)
                qty_med = _quantile(qtys_s, 0.5) if qtys_s else None
                qty_p10 = _quantile(qtys_s, 0.10) if qtys_s else None
                qty_p90 = _quantile(qtys_s, 0.90) if qtys_s else None

                qty_hhi = None
                qty_top1_share = None
                if qty_by_seller:
                    qcnts = list(qty_by_seller.values())
                    qty_hhi = _hhi_from_counts(qcnts)
                    qty_top1_share = _top1_share_from_counts(qcnts)

                # --- FLAGS (rules-first) ---
                flags: List[str] = []
                # Confidence
                if len(prices) < min_n or len(kept) < min_n:
                    flags.append("LOW_CONFIDENCE")
                # Dumping pressure: cheap tail vs median + many sellers
                if p_med and p10 and unique_sellers >= 10:
                    if p10 < 0.75 * p_med and (out_rate >= 0.10 or len(prices) >= 12):
                        flags.append("DUMPING_PRESSURE")
                # Monopoly danger: seller concentration (counts)
                if top1_share is not None and top1_share >= 0.35:
                    flags.append("MONOPOLY_DANGER")
                if hhi is not None and hhi >= 0.25:
                    if "MONOPOLY_DANGER" not in flags:
                        flags.append("MONOPOLY_DANGER")

                global_flags.extend(flags)

                out_clusters.append({
                    "cluster": cl,
                    "n_kept": len(kept),
                    "price": {
                        "n": len(prices),
                        "median": round(p_med, 2) if p_med is not None else None,
                        "p10": round(p10, 2) if p10 is not None else None,
                        "p90": round(p90, 2) if p90 is not None else None,
                        "trimmed_median_10": round(p_trim_med, 2) if p_trim_med is not None else None,
                        "outliers_low_n": int(low_out),
                        "outliers_high_n": int(high_out),
                        "outlier_rate": float(out_rate),
                    },
                    "stock_proxy": {
                        "n": len(qtys),
                        "median": float(qty_med) if qty_med is not None else None,
                        "p10": float(qty_p10) if qty_p10 is not None else None,
                        "p90": float(qty_p90) if qty_p90 is not None else None,
                        "seller_top1_share_qty": float(qty_top1_share) if qty_top1_share is not None else None,
                        "seller_hhi_qty": float(qty_hhi) if qty_hhi is not None else None,
                    },
                    "seller_structure": {
                        "unique_sellers": int(unique_sellers),
                        "top1_share_count": float(top1_share) if top1_share is not None else None,
                        "hhi_count": float(hhi) if hhi is not None else None,
                        "total_listings": len(kept),
                    },
                    "risk_flags": sorted(set(flags)),
                })

            append_jsonl(out_path, {
                "meta": {
                    "schema_version": "2.0",
                    "run_id": run_id,
                    "nm_id": own_nm,
                    "vendor_code": str(meta.get("vendor_code") or ""),
                    "ts": _iso_utc(),
                    "stage": "J",
                },
                "supply_structure": {"clusters": out_clusters},
                "notes": notes_in,
                "risk_flags": sorted(set(global_flags)),
            })

            processed += 1

    print(f"[J] processed: {processed}")
    return out_path

# =========================
# Stage K: CLUSTER VERDICTS (WB Necromancer v2)
# Reads:  run_manifest.json, market_pulse.jsonl, supply_structure.jsonl, relevance.jsonl
# Writes: cluster_verdicts.jsonl
# Network: NO (LOCAL)
# =========================

# from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# --- minimal helpers ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _load_jsonl_map(path: Path, key_field: str = "nm_id") -> Dict[str, dict]:
    m: Dict[str, dict] = {}
    if not path.exists():
        return m
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            j = json.loads(line)
            nm = str((j.get("meta") or {}).get(key_field) or "")
            if nm:
                m[nm] = j
    return m


def _pulse_to_status(cluster_row: dict) -> Tuple[str, str, List[str]]:
    """
    English technical:
    Rules-based ALIVE/SLOW/DEAD from review velocity.
    """
    flags: List[str] = []

    imt_n = int(cluster_row.get("imt_n") or 0)
    imt_ok = int(cluster_row.get("imt_with_data_n") or 0)

    r30_med = cluster_row.get("recent_30_median")
    r90_med = cluster_row.get("recent_90_median")
    dsl_med = cluster_row.get("days_since_last_median")
    dsl_min = cluster_row.get("days_since_last_min")

    # Confidence
    if imt_ok >= 10:
        conf = "HIGH"
    elif imt_ok >= 6:
        conf = "MEDIUM"
    else:
        conf = "LOW"
        flags.append("LOW_CONFIDENCE")

    # No data -> DEAD (low confidence)
    if imt_ok == 0:
        return "DEAD", conf, flags + ["NO_REVIEW_DATA"]

    # Normalize numbers (some may be None/float)
    try:
        r30 = float(r30_med) if r30_med is not None else 0.0
    except Exception:
        r30 = 0.0
    try:
        r90 = float(r90_med) if r90_med is not None else 0.0
    except Exception:
        r90 = 0.0

    # Use min/median freshness if present
    dmin = int(dsl_min) if dsl_min is not None else None
    dmed = int(dsl_med) if dsl_med is not None else None

    # Strong alive signals
    if r30 >= 1.0:
        return "ALIVE", conf, flags
    if dmin is not None and dmin <= 21:
        return "ALIVE", conf, flags
    if r90 >= 3.0 and (dmed is None or dmed <= 45):
        return "ALIVE", conf, flags

    # Slow market signals
    if r90 >= 1.0 and (dmed is None or dmed <= 90):
        return "SLOW", conf, flags
    if dmin is not None and dmin <= 120 and r90 > 0:
        return "SLOW", conf, flags

    # Otherwise dead
    return "DEAD", conf, flags


def stage_K_cluster_verdicts(
    out_dir: Path,
    *,
    resume: bool = False,
) -> Path:
    """
    Stage K:
    - merges Market Pulse (I) + Supply/Structure (J) + notes (H)
    - emits per cluster status: ALIVE/SLOW/DEAD + confidence
    """
    manifest_path = out_dir / "run_manifest.json"
    mp_path = out_dir / "market_pulse.jsonl"
    ss_path = out_dir / "supply_structure.jsonl"
    rel_path = out_dir / "relevance.jsonl"

    if not manifest_path.exists():
        raise FileNotFoundError("run_manifest.json not found; run Stage A")
    if not mp_path.exists():
        raise FileNotFoundError("market_pulse.jsonl not found; run Stage I")
    if not ss_path.exists():
        raise FileNotFoundError("supply_structure.jsonl not found; run Stage J")
    if not rel_path.exists():
        raise FileNotFoundError("relevance.jsonl not found; run Stage H")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    mp_map = _load_jsonl_map(mp_path, "nm_id")
    ss_map = _load_jsonl_map(ss_path, "nm_id")
    rel_map = _load_jsonl_map(rel_path, "nm_id")

    out_path = out_dir / "cluster_verdicts.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    processed = 0
    for own_nm, mp in mp_map.items():
        if not own_nm or own_nm in done:
            continue

        ss = ss_map.get(own_nm) or {}
        rel = rel_map.get(own_nm) or {}

        vendor_code = str(((mp.get("meta") or {}).get("vendor_code")) or
                          ((ss.get("meta") or {}).get("vendor_code")) or
                          ((rel.get("meta") or {}).get("vendor_code")) or "")

        mp_clusters = ((mp.get("market_pulse") or {}).get("clusters") or [])
        ss_clusters = (((ss.get("supply_structure") or {}).get("clusters")) or [])

        notes = []
        notes.extend(list(rel.get("notes") or []))
        notes.extend(list((mp.get("notes") or [])))
        notes.extend(list((ss.get("notes") or [])))

        out_clusters: List[dict] = []
        global_flags: List[str] = []

        for cl in ("phone", "type"):
            mp_row = next((x for x in mp_clusters if str(x.get("cluster") or "") == cl), {}) or {}
            ss_row = next((x for x in ss_clusters if str(x.get("cluster") or "") == cl), {}) or {}

            status, conf, pulse_flags = _pulse_to_status(mp_row)
            risk_flags = sorted(set((ss_row.get("risk_flags") or []) + pulse_flags))

            # If supply structure itself is low confidence, downgrade confidence one notch
            if "LOW_CONFIDENCE" in risk_flags and conf == "HIGH":
                conf = "MEDIUM"
            elif "LOW_CONFIDENCE" in risk_flags and conf == "MEDIUM":
                conf = "LOW"

            # Attach risk flags up
            global_flags.extend(risk_flags)

            out_clusters.append({
                "cluster": cl,
                "status": status,
                "confidence": conf,
                "facts": {
                    "market_pulse": {
                        "imt_n": mp_row.get("imt_n"),
                        "imt_with_data_n": mp_row.get("imt_with_data_n"),
                        "recent_30_sum": mp_row.get("recent_30_sum"),
                        "recent_90_sum": mp_row.get("recent_90_sum"),
                        "recent_30_median": mp_row.get("recent_30_median"),
                        "recent_90_median": mp_row.get("recent_90_median"),
                        "days_since_last_median": mp_row.get("days_since_last_median"),
                        "days_since_last_min": mp_row.get("days_since_last_min"),
                    },
                    "supply_structure": {
                        "price": (ss_row.get("price") or {}),
                        "seller_structure": (ss_row.get("seller_structure") or {}),
                        "stock_proxy": (ss_row.get("stock_proxy") or {}),
                    }
                },
                "risk_flags": risk_flags,
            })

        append_jsonl(out_path, {
            "meta": {
                "schema_version": "2.0",
                "run_id": run_id,
                "nm_id": own_nm,
                "vendor_code": vendor_code,
                "ts": _iso_utc(),
                "stage": "K",
            },
            "cluster_verdicts": out_clusters,
            "risk_flags": sorted(set(global_flags)),
            "notes": notes,
        })

        processed += 1

    print(f"[K] processed: {processed}")
    return out_path

# =========================
# Stage L: FINAL DECISION (WB Necromancer v2)
# Reads:  run_manifest.json, intent.jsonl, cluster_verdicts.jsonl
# Writes: decisions.jsonl
# Network: LOCAL (optional LLM for phrasing)
# =========================

# from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# --- minimal helpers ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def read_jsonl_nm_ids(path: Path) -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                nm = str((j.get("meta") or {}).get("nm_id") or "")
                if nm:
                    done.add(nm)
            except Exception:
                continue
    return done

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def _load_jsonl_map(path: Path, key_field: str = "nm_id") -> Dict[str, dict]:
    m: Dict[str, dict] = {}
    if not path.exists():
        return m
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
            except Exception:
                continue
            nm = str((j.get("meta") or {}).get(key_field) or "")
            if nm:
                m[nm] = j
    return m


def _get_cluster_row(verdict_rec: dict, cluster: str) -> dict:
    rows = verdict_rec.get("cluster_verdicts") or []
    for r in rows:
        if str(r.get("cluster") or "") == cluster:
            return r
    return {}

def _get_karma(intent_rec: dict) -> Tuple[Optional[float], Optional[int]]:
    k = (intent_rec.get("intent") or {}).get("karma") or {}
    rating = k.get("rating")
    feedbacks = k.get("feedbacks")
    try:
        rating = float(rating) if rating is not None else None
    except Exception:
        rating = None
    try:
        feedbacks = int(float(feedbacks)) if feedbacks is not None else None
    except Exception:
        feedbacks = None
    return rating, feedbacks

def _karma_is_toxic(
    rating: Optional[float],
    feedbacks: Optional[int],
    *,
    min_fb: int,
    min_rating: float
) -> bool:
    # “токсична” = низкий рейтинг при достаточном числе отзывов
    if rating is None or feedbacks is None:
        return False
    return feedbacks >= int(min_fb) and rating < float(min_rating)

def _backlog_from_issues(issues: List[str], *, decision: str) -> List[str]:
    tasks: List[str] = []
    issues = [str(x) for x in (issues or [])]

    if "MISSING_PHONE_MODEL" in issues:
        tasks.append("Уточнить и зафиксировать модель телефона (в названии, характеристиках, описании).")
    if "NO_TPU_SIGNAL" in issues:
        tasks.append("Явно указать материал TPU и синонимы (силикон, термополиуретан) в тексте и атрибутах.")
    if "NO_POCKET_SIGNAL" in issues:
        tasks.append("Явно указать карман или картхолдер (карман для карт, слот под карту) в тексте и на фото.")

    if decision in ("REVIVE_REWORK", "CLONE_NEW_CARD"):
        tasks.extend([
            "Пересобрать SEO: заголовок до 60 символов, ключи без спама, заполнить характеристики под фильтры.",
            "Проверить фото: добавить кадр с карманом для карты и материалом, привести к единому стилю.",
        ])

    if decision == "CLONE_NEW_CARD":
        tasks.append("Создать новую карточку и перенести лучший контент без токсичной истории отзывов.")

    return tasks[:12]


def stage_L_final_decision(
    out_dir: Path,
    *,
    resume: bool = False,
    use_llm_l: bool = False,
    llm_provider: str = "openrouter",
    llm_model: str = "openai/gpt-4o-mini",
    api_key: str = "",
    llm_base_url: str = "",
    llm_timeout: int = 60,
    llm_max_tokens: int = 1200,
    llm_temperature: float = 0.2,
) -> Path:
    """
    Stage L:
    - Applies v2 decision matrix (rules-first)
    - Optional: LLM phrasing based on FACTS (must not change decision)
    Output enum: DROP | REVIVE_FAST | REVIVE_REWORK | CLONE_NEW_CARD
    """
    manifest_path = out_dir / "run_manifest.json"
    intent_path = out_dir / "intent.jsonl"
    cv_path = out_dir / "cluster_verdicts.jsonl"

    if not manifest_path.exists():
        raise FileNotFoundError("run_manifest.json not found; run Stage A")
    if not intent_path.exists():
        raise FileNotFoundError("intent.jsonl not found; run Stage C")
    if not cv_path.exists():
        raise FileNotFoundError("cluster_verdicts.jsonl not found; run Stage K")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    # Configurable karma thresholds (defaults match v2 intent)
    cfg = manifest.get("config") or {}
    karma_cfg = cfg.get("karma") or {}
    karma_min_fb = int(karma_cfg.get("min_feedbacks") or 25)
    karma_toxic_rating_lt = float(karma_cfg.get("toxic_rating_lt") or 4.4)

    intent_map = _load_jsonl_map(intent_path, "nm_id")
    cv_map = _load_jsonl_map(cv_path, "nm_id")

    out_path = out_dir / "decisions.jsonl"
    done = read_jsonl_nm_ids(out_path) if resume else set()

    processed = 0

    for own_nm, cv in cv_map.items():
        if not own_nm or own_nm in done:
            continue

        intent_rec = intent_map.get(own_nm) or {}
        vendor_code = str(
            ((cv.get("meta") or {}).get("vendor_code")) or
            ((intent_rec.get("meta") or {}).get("vendor_code")) or
            ""
        )

        phone_row = _get_cluster_row(cv, "phone")
        type_row = _get_cluster_row(cv, "type")

        phone_status = str(phone_row.get("status") or "DEAD")
        type_status = str(type_row.get("status") or "DEAD")

        rating, feedbacks = _get_karma(intent_rec)
        issues = list(((intent_rec.get("intent") or {}).get("issues")) or [])
        extracted = ((intent_rec.get("intent") or {}).get("extracted")) or {}

        toxic = _karma_is_toxic(
            rating, feedbacks,
            min_fb=karma_min_fb,
            min_rating=karma_toxic_rating_lt
        )

        risk_flags = sorted(set(
            (cv.get("risk_flags") or []) +
            (phone_row.get("risk_flags") or []) +
            (type_row.get("risk_flags") or [])
        ))
        notes = list(cv.get("notes") or [])

        decision = "DROP"
        alt_strategy_ru: Optional[str] = None

        # =========================
        # Decision matrix v2
        # =========================
        if phone_status == "DEAD":
            decision = "DROP"
            risk_flags = sorted(set(risk_flags + ["PHONE_MARKET_DEAD"]))
        else:
            # phone is ALIVE or SLOW
            if type_status == "DEAD":
                decision = "DROP"
                risk_flags = sorted(set(risk_flags + ["TYPE_MARKET_DEAD"]))
                alt_strategy_ru = (
                    "Модель телефона выглядит живой, но тип «TPU + карман для карты» не подтверждён рынком. "
                    "Если хочешь работать по модели, рассмотри другие типы чехлов (обычный TPU, противоударный, книжка)."
                )
            else:
                # type is ALIVE or SLOW
                if toxic:
                    decision = "CLONE_NEW_CARD"
                    risk_flags = sorted(set(risk_flags + ["TOXIC_KARMA"]))
                else:
                    content_debt = False

                    if "MISSING_PHONE_MODEL" in issues:
                        content_debt = True
                    if "NO_TPU_SIGNAL" in issues or "NO_POCKET_SIGNAL" in issues:
                        content_debt = True

                    # risk flags that usually force “думать, а не жать кнопку”
                    if "LOW_CONFIDENCE" in risk_flags:
                        content_debt = True
                    if "DUMPING_PRESSURE" in risk_flags or "MONOPOLY_DANGER" in risk_flags:
                        content_debt = True

                    if type_status == "SLOW":
                        content_debt = True
                        risk_flags = sorted(set(risk_flags + ["SLOW_DEMAND"]))

                    decision = "REVIVE_REWORK" if content_debt else "REVIVE_FAST"

        backlog = _backlog_from_issues(issues, decision=decision)

        # =========================
        # Deterministic rationale
        # =========================
        phone_conf = str(phone_row.get("confidence") or "")
        type_conf = str(type_row.get("confidence") or "")

        karma_part = ""
        if rating is not None and feedbacks is not None:
            karma_part = f"Карма карточки: рейтинг {rating:.2f}, отзывов {feedbacks}."
        elif rating is not None:
            karma_part = f"Карма карточки: рейтинг {rating:.2f}."
        elif feedbacks is not None:
            karma_part = f"Карма карточки: отзывов {feedbacks}."

        rationale_ru = (
            f"Рынок модели (phone) = {phone_status} (conf={phone_conf}); "
            f"рынок типа TPU+карман (type) = {type_status} (conf={type_conf}). "
            f"{karma_part} "
        ).strip()

        if decision == "DROP":
            if phone_status == "DEAD":
                rationale_ru += " Общий рынок чехлов на модель не выглядит живым: оживление SKU не окупится."
            else:
                rationale_ru += " Модель в целом жива, но спрос на «TPU + карман» не подтверждён: обычно это повод не тратить ресурсы."
        elif decision == "CLONE_NEW_CARD":
            rationale_ru += " Типовой рынок подтверждён, но история карточки токсична: клонирование в новую карточку рациональнее."
        elif decision == "REVIVE_FAST":
            rationale_ru += " Оба рынка выглядят живыми и без крупных рисков: можно оживлять быстро."
        else:
            rationale_ru += " Рынки не мёртвые, но есть риски или контент-долг: нужна переработка перед оживлением."

        # =========================
        # Optional LLM phrasing (FACTS-only)
        # =========================
        llm_used = False
        llm_debug: Dict[str, Any] = {}

        if use_llm_l:
            # Expect call_llm_json() to exist in the final script (borrow from wb_revive.py).
            if "call_llm_json" not in globals():
                raise RuntimeError("use_llm_l=True but call_llm_json() is not defined in this script.")

            facts = {
                "decision_fixed": decision,
                "phone_market": phone_row,
                "type_market": type_row,
                "karma": {"rating": rating, "feedbacks": feedbacks, "toxic": toxic},
                "risk_flags": risk_flags,
                "issues": issues,
                "extracted": extracted,
                "alt_strategy_ru": alt_strategy_ru,
            }

            messages = [
                {"role": "system", "content": (
                    "You are a strict analyst. Output JSON only. "
                    "Do not invent numbers. Use only FACTS. "
                    "You must not change decision_fixed."
                )},
                {"role": "user", "content": (
                    "FACTS (JSON):\n"
                    f"{json.dumps(facts, ensure_ascii=False)}\n\n"
                    "Return JSON object with keys: "
                    "rationale_ru (string), backlog_ru (array of strings), risk_flags (array of EN UPPER_SNAKE_CASE), "
                    "alt_strategy_ru (string or null). Keep it concise."
                )},
            ]

            obj, dbg = call_llm_json(
                provider=llm_provider,
                model=llm_model,
                api_key=api_key,
                messages=messages,
                base_url=llm_base_url,
                timeout_sec=llm_timeout,
                max_tokens=llm_max_tokens,
                temperature=llm_temperature,
                force_json=True,
            )
            llm_used = True
            llm_debug = dbg or {}

            if isinstance(obj, dict):
                rr = obj.get("rationale_ru")
                if isinstance(rr, str) and rr.strip():
                    rationale_ru = rr.strip()

                bl = obj.get("backlog_ru")
                if isinstance(bl, list) and bl:
                    backlog = [str(x) for x in bl if str(x).strip()][:12]

                rf = obj.get("risk_flags")
                if isinstance(rf, list) and rf:
                    risk_flags = sorted(set(risk_flags + [str(x) for x in rf if str(x).strip()]))

                ar = obj.get("alt_strategy_ru")
                if isinstance(ar, str) and ar.strip():
                    alt_strategy_ru = ar.strip()

        append_jsonl(out_path, {
            "meta": {
                "schema_version": "2.0",
                "run_id": run_id,
                "nm_id": str(own_nm),
                "vendor_code": vendor_code,
                "ts": _iso_utc(),
                "stage": "L",
            },
            "decision": {
                "final": decision,  # DROP|REVIVE_FAST|REVIVE_REWORK|CLONE_NEW_CARD
                "phone_status": phone_status,
                "type_status": type_status,
                "confidence": {
                    "phone": phone_row.get("confidence"),
                    "type": type_row.get("confidence"),
                },
                "karma": {
                    "rating": rating,
                    "feedbacks": feedbacks,
                    "toxic": toxic,
                    "toxic_rule": {"min_feedbacks": karma_min_fb, "rating_lt": karma_toxic_rating_lt},
                },
                "risk_flags": risk_flags,
                "alt_strategy_ru": alt_strategy_ru,
            },
            "rationale_ru": rationale_ru,
            "backlog_ru": backlog,
            "evidence": {
                "cluster_verdicts": cv.get("cluster_verdicts"),
                "facts_phone": phone_row.get("facts"),
                "facts_type": type_row.get("facts"),
                "issues": issues,
                "notes": notes,
            },
            "llm": {
                "used": bool(llm_used),
                "provider": llm_provider if llm_used else "",
                "model": llm_model if llm_used else "",
                "debug": llm_debug,
            }
        })

        processed += 1

    return out_path

# =========================
# Stage M: REPORTS (WB Necromancer v2)
# Reads:  run_manifest.json, decisions.jsonl
# Writes: WB_REVIVE_REPORT.xlsx, WB_REVIVE_REPORT.html, (optional) exec_summary.json
# Network: LOCAL (optional LLM)
# =========================

# from __future__ import annotations

import json
import html
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# --- minimal helpers ---
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def _iso_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def _load_jsonl_list(path: Path) -> List[dict]:
    out = []
    if not path.exists():
        return out
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except Exception:
                continue
    return out


# --- report helpers ---
def _fmt_flags(flags: Any) -> str:
    if isinstance(flags, list):
        return ", ".join(str(x) for x in flags if str(x).strip())
    if isinstance(flags, str):
        return flags
    return ""

def _fmt_list(lines: Any, sep: str = " | ") -> str:
    if isinstance(lines, list):
        return sep.join(str(x).strip() for x in lines if str(x).strip())
    if isinstance(lines, str):
        return lines.strip()
    return ""

def _get_cluster_fact(dec: dict, cluster: str) -> dict:
    # decisions.jsonl: evidence.facts_phone/type stored as dict with market_pulse + supply_structure
    ev = (dec.get("evidence") or {})
    if cluster == "phone":
        return (ev.get("facts_phone") or {}) if isinstance(ev.get("facts_phone"), dict) else {}
    if cluster == "type":
        return (ev.get("facts_type") or {}) if isinstance(ev.get("facts_type"), dict) else {}
    return {}

def _safe_num(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        return float(x)
    except Exception:
        return None

def _pick(d: dict, path: List[str], default=None):
    cur = d
    for k in path:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k)
    return cur if cur is not None else default


# --- XLSX styling ---
HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # dark gray
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(vertical="top", wrap_text=True)
THIN_FILL_DROP = PatternFill("solid", fgColor="FEE2E2")     # light red
THIN_FILL_FAST = PatternFill("solid", fgColor="DCFCE7")     # light green
THIN_FILL_REWORK = PatternFill("solid", fgColor="FEF9C3")   # light yellow
THIN_FILL_CLONE = PatternFill("solid", fgColor="E0E7FF")    # light indigo


def _apply_header(ws: Worksheet, headers: List[str]) -> None:
    ws.append(headers)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _autosize(ws: Worksheet, max_width: int = 62) -> None:
    # crude but effective: measure cell string lengths
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            s = "" if v is None else str(v)
            widths[i] = max(widths.get(i, 8), min(max_width, len(s) + 2))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def _fill_decision_cell(cell, decision: str) -> None:
    d = (decision or "").upper().strip()
    if d == "DROP":
        cell.fill = THIN_FILL_DROP
    elif d == "REVIVE_FAST":
        cell.fill = THIN_FILL_FAST
    elif d == "REVIVE_REWORK":
        cell.fill = THIN_FILL_REWORK
    elif d == "CLONE_NEW_CARD":
        cell.fill = THIN_FILL_CLONE


def _write_xlsx(report_path: Path, rows: List[dict], meta: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "REPORT"

    headers = [
        "nm_id", "vendor_code", "final_decision",
        "phone_status", "phone_conf",
        "type_status", "type_conf",
        "karma_rating", "karma_feedbacks", "karma_toxic",
        "risk_flags",
        # Market pulse (phone)
        "phone_recent30_med", "phone_recent90_med", "phone_days_since_last_med",
        # Supply (phone)
        "phone_price_med", "phone_price_p10", "phone_price_p90", "phone_unique_sellers",
        # Market pulse (type)
        "type_recent30_med", "type_recent90_med", "type_days_since_last_med",
        # Supply (type)
        "type_price_med", "type_price_p10", "type_price_p90", "type_unique_sellers",
        # Text
        "rationale_ru", "backlog_ru", "alt_strategy_ru",
    ]
    _apply_header(ws, headers)

    for r in rows:
        dec = r.get("decision") or {}
        final_dec = _pick(dec, ["final"], "")

        ws.append([
            r.get("nm_id", ""),
            r.get("vendor_code", ""),
            final_dec,

            r.get("phone_status", ""),
            r.get("phone_conf", ""),

            r.get("type_status", ""),
            r.get("type_conf", ""),

            r.get("karma_rating", None),
            r.get("karma_feedbacks", None),
            r.get("karma_toxic", None),

            r.get("risk_flags", ""),

            r.get("phone_recent30_med", None),
            r.get("phone_recent90_med", None),
            r.get("phone_days_since_last_med", None),

            r.get("phone_price_med", None),
            r.get("phone_price_p10", None),
            r.get("phone_price_p90", None),
            r.get("phone_unique_sellers", None),

            r.get("type_recent30_med", None),
            r.get("type_recent90_med", None),
            r.get("type_days_since_last_med", None),

            r.get("type_price_med", None),
            r.get("type_price_p10", None),
            r.get("type_price_p90", None),
            r.get("type_unique_sellers", None),

            r.get("rationale_ru", ""),
            r.get("backlog_ru", ""),
            r.get("alt_strategy_ru", ""),
        ])

    # style columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = LEFT_WRAP

    # highlight decision column (C)
    for i in range(2, ws.max_row + 1):
        c = ws.cell(row=i, column=3)
        _fill_decision_cell(c, str(c.value or ""))

    # add META sheet
    ws2 = wb.create_sheet("META")
    ws2.append(["key", "value"])
    ws2["A1"].fill = HEADER_FILL
    ws2["A1"].font = HEADER_FONT
    ws2["B1"].fill = HEADER_FILL
    ws2["B1"].font = HEADER_FONT
    for k, v in (meta or {}).items():
        ws2.append([str(k), json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)])

    _autosize(ws)
    _autosize(ws2, max_width=80)

    ensure_dir(report_path.parent)
    wb.save(report_path)


# --- HTML report ---
def _esc(s: Any) -> str:
    return html.escape("" if s is None else str(s))

def _write_html(html_path: Path, rows: List[dict], meta: dict, exec_summary: Optional[dict] = None) -> None:
    # Minimal interactive-ish table (no external libs)
    head = f"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>WB Revival Report</title>
<style>
body {{ font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial; padding: 16px; }}
h1 {{ margin: 0 0 8px; }}
small {{ color: #555; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 12px; }}
th, td {{ border: 1px solid #ddd; padding: 8px; vertical-align: top; }}
th {{ background: #111827; color: #fff; position: sticky; top: 0; z-index: 2; }}
tr:nth-child(even) {{ background: #fafafa; }}
.badge {{ padding: 2px 8px; border-radius: 10px; font-weight: 600; display: inline-block; }}
.b-drop {{ background: #fee2e2; }}
.b-fast {{ background: #dcfce7; }}
.b-rework {{ background: #fef9c3; }}
.b-clone {{ background: #e0e7ff; }}
.mono {{ font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; }}
details > summary {{ cursor: pointer; }}
</style>
</head>
<body>
<h1>WB Revival v2 Necromancer</h1>
<small class="mono">generated_at_utc={_esc(_iso_utc())} | run_id={_esc((meta or {}).get("run_id",""))}</small>
"""

    if exec_summary and isinstance(exec_summary, dict):
        head += "<h2>Exec summary</h2>\n"
        head += "<pre class='mono' style='white-space:pre-wrap;border:1px solid #ddd;padding:12px;background:#f9fafb;'>"
        head += _esc(json.dumps(exec_summary, ensure_ascii=False, indent=2))
        head += "</pre>\n"

    head += "<h2>Table</h2>\n"
    head += "<p><small>Два рынка: phone=общий спрос по модели, type=TPU+карман. Решение берётся из Stage L.</small></p>\n"
    head += "<table><thead><tr>"
    cols = [
        "nm_id","vendor_code","final_decision",
        "phone_status","type_status",
        "karma_rating","karma_feedbacks","karma_toxic",
        "risk_flags",
        "phone_recent30_med","phone_recent90_med","phone_days_since_last_med","phone_price_med","phone_unique_sellers",
        "type_recent30_med","type_recent90_med","type_days_since_last_med","type_price_med","type_unique_sellers",
        "rationale_ru",
        "backlog_ru",
        "alt_strategy_ru",
    ]
    for c in cols:
        head += f"<th>{_esc(c)}</th>"
    head += "</tr></thead><tbody>\n"

    def badge(dec: str) -> str:
        d = (dec or "").upper()
        cls = "badge "
        if d == "DROP":
            cls += "b-drop"
        elif d == "REVIVE_FAST":
            cls += "b-fast"
        elif d == "REVIVE_REWORK":
            cls += "b-rework"
        elif d == "CLONE_NEW_CARD":
            cls += "b-clone"
        else:
            cls += "mono"
        return f"<span class='{cls}'>{_esc(d)}</span>"

    rows_html = ""
    for r in rows:
        rows_html += "<tr>"
        for c in cols:
            v = r.get(c, "")
            if c == "final_decision":
                v = badge(str(v))
                rows_html += f"<td>{v}</td>"
            else:
                rows_html += f"<td>{_esc(v)}</td>"
        rows_html += "</tr>\n"

    tail = "</tbody></table>\n</body></html>"
    ensure_dir(html_path.parent)
    html_path.write_text(head + rows_html + tail, encoding="utf-8")


def stage_M_reports(
    out_dir: Path,
    *,
    report_xlsx_name: str = "WB_REVIVE_REPORT.xlsx",
    report_html_name: str = "WB_REVIVE_REPORT.html",
    resume: bool = False,  # kept for menu symmetry; Stage M is cheap, usually overwrite OK
    use_llm_m: bool = False,
    llm_provider: str = "openrouter",
    llm_model: str = "openai/gpt-4o-mini",
    api_key: str = "",
    llm_base_url: str = "",
    llm_timeout: int = 60,
    llm_max_tokens: int = 1400,
    llm_temperature: float = 0.2,
) -> Tuple[Path, Path]:
    """
    Stage M:
    - Build XLSX + HTML reports from decisions.jsonl
    - Optional: exec summary JSON using LLM (facts-only)
    """
    manifest_path = out_dir / "run_manifest.json"
    dec_path = out_dir / "decisions.jsonl"
    if not manifest_path.exists():
        raise FileNotFoundError("run_manifest.json not found; run Stage A")
    if not dec_path.exists():
        raise FileNotFoundError("decisions.jsonl not found; run Stage L")

    manifest = read_json(manifest_path)
    run_id = str(manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("Bad manifest: missing run_id")

    decs = _load_jsonl_list(dec_path)

    # Transform to flat rows with two-market facts
    rows: List[dict] = []
    counts = {"DROP": 0, "REVIVE_FAST": 0, "REVIVE_REWORK": 0, "CLONE_NEW_CARD": 0}

    for d in decs:
        meta = d.get("meta") or {}
        nm_id = str(meta.get("nm_id") or "")
        vendor_code = str(meta.get("vendor_code") or "")

        decision = d.get("decision") or {}
        final_dec = str(decision.get("final") or "")
        counts[final_dec] = counts.get(final_dec, 0) + 1

        phone_fact = _get_cluster_fact(d, "phone")
        type_fact = _get_cluster_fact(d, "type")

        # market pulse fields
        ph_mp = (phone_fact.get("market_pulse") or {}) if isinstance(phone_fact.get("market_pulse"), dict) else {}
        ty_mp = (type_fact.get("market_pulse") or {}) if isinstance(type_fact.get("market_pulse"), dict) else {}

        # supply fields
        ph_ss = (phone_fact.get("supply_structure") or {}) if isinstance(phone_fact.get("supply_structure"), dict) else {}
        ty_ss = (type_fact.get("supply_structure") or {}) if isinstance(type_fact.get("supply_structure"), dict) else {}

        ph_price = (ph_ss.get("price") or {}) if isinstance(ph_ss.get("price"), dict) else {}
        ty_price = (ty_ss.get("price") or {}) if isinstance(ty_ss.get("price"), dict) else {}

        ph_sellers = (ph_ss.get("seller_structure") or {}) if isinstance(ph_ss.get("seller_structure"), dict) else {}
        ty_sellers = (ty_ss.get("seller_structure") or {}) if isinstance(ty_ss.get("seller_structure"), dict) else {}

        karma = (decision.get("karma") or {}) if isinstance(decision.get("karma"), dict) else {}

        row = {
            "nm_id": nm_id,
            "vendor_code": vendor_code,

            "decision": decision,
            "final_decision": final_dec,

            "phone_status": decision.get("phone_status"),
            "phone_conf": _pick(decision, ["confidence", "phone"], ""),
            "type_status": decision.get("type_status"),
            "type_conf": _pick(decision, ["confidence", "type"], ""),

            "karma_rating": karma.get("rating"),
            "karma_feedbacks": karma.get("feedbacks"),
            "karma_toxic": karma.get("toxic"),

            "risk_flags": _fmt_flags(decision.get("risk_flags")),

            "phone_recent30_med": ph_mp.get("recent_30_median"),
            "phone_recent90_med": ph_mp.get("recent_90_median"),
            "phone_days_since_last_med": ph_mp.get("days_since_last_median"),

            "phone_price_med": ph_price.get("median"),
            "phone_price_p10": ph_price.get("p10"),
            "phone_price_p90": ph_price.get("p90"),
            "phone_unique_sellers": ph_sellers.get("unique_sellers"),

            "type_recent30_med": ty_mp.get("recent_30_median"),
            "type_recent90_med": ty_mp.get("recent_90_median"),
            "type_days_since_last_med": ty_mp.get("days_since_last_median"),

            "type_price_med": ty_price.get("median"),
            "type_price_p10": ty_price.get("p10"),
            "type_price_p90": ty_price.get("p90"),
            "type_unique_sellers": ty_sellers.get("unique_sellers"),

            "rationale_ru": d.get("rationale_ru") or "",
            "backlog_ru": _fmt_list(d.get("backlog_ru"), sep=" | "),
            "alt_strategy_ru": _pick(decision, ["alt_strategy_ru"], "") or "",
        }
        rows.append(row)

    # Optional exec summary (facts-only)
    exec_summary = None
    if use_llm_m:
        if "call_llm_json" not in globals():
            raise RuntimeError("use_llm_m=True but call_llm_json() is not defined in this script.")
        facts = {
            "run_id": run_id,
            "total_sku": len(rows),
            "decision_counts": counts,
            "notes": "Summarize briefly. No invented numbers.",
        }
        messages = [
            {"role": "system", "content": "Return JSON only. Use only provided FACTS. No invention."},
            {"role": "user", "content": (
                "FACTS:\n" + json.dumps(facts, ensure_ascii=False) + "\n\n"
                "Return JSON with keys: headline_ru, bullets_ru (array), risks_ru (array), next_steps_ru (array)."
            )},
        ]
        obj, _dbg = call_llm_json(
            provider=llm_provider,
            model=llm_model,
            api_key=api_key,
            messages=messages,
            base_url=llm_base_url,
            timeout_sec=llm_timeout,
            max_tokens=llm_max_tokens,
            temperature=llm_temperature,
            force_json=True,
        )
        exec_summary = obj if isinstance(obj, dict) else None
        if exec_summary is not None:
            (out_dir / "exec_summary.json").write_text(json.dumps(exec_summary, ensure_ascii=False, indent=2), encoding="utf-8")

    report_xlsx = out_dir / report_xlsx_name
    report_html = out_dir / report_html_name

    meta_out = {
        "run_id": run_id,
        "generated_at_utc": _iso_utc(),
        "counts": counts,
        "source": {"decisions": str(dec_path)},
    }

    _write_xlsx(report_xlsx, rows, meta_out)
    _write_html(report_html, rows, meta_out, exec_summary=exec_summary)

    print(f"[M] wrote: {report_xlsx}")
    print(f"[M] wrote: {report_html}")
    return report_xlsx, report_html

  ###Если хочешь прям “как в wb_revive.py”, то в раннере сделай:

 #print_stage_banner("M", args, out_dir=str(out_dir))

 #вызов stage_M_reports(out_dir, use_llm_m=args.use_llm_m, ...)

 


# =========================
# CLI Runner (rebuilt)
# =========================

import argparse

STAGE_ORDER = ["A","B","C","D","E","F","G","H","I","J","K","L","M"]

STAGE_INFO = {
    "A": {"title":"INPUT + Manifest", "network":"LOCAL", "llm_capable": False},
    "B": {"title":"OWN FETCH (WB)", "network":"WB", "llm_capable": False},
    "C": {"title":"INTENT EXTRACT", "network":"LOCAL", "llm_capable": False},
    "D": {"title":"QUERY BUILD (rules-first)", "network":"LOCAL", "llm_capable": False},  # llm hook not wired here yet
    "E": {"title":"SERP SNAPSHOT + VALIDATION", "network":"WB", "llm_capable": False},
    "F": {"title":"COMPETITOR POOL", "network":"WB", "llm_capable": False},
    "G": {"title":"COMPETITOR LITE FETCH", "network":"WB", "llm_capable": False},
    "H": {"title":"RELEVANCE FILTER (rules-only)", "network":"LOCAL", "llm_capable": False},  # v1: rules-only
    "I": {"title":"MARKET PULSE (reviews)", "network":"WB", "llm_capable": False},
    "J": {"title":"SUPPLY/STRUCTURE", "network":"LOCAL", "llm_capable": False},
    "K": {"title":"CLUSTER VERDICTS", "network":"LOCAL", "llm_capable": False},
    "L": {"title":"FINAL DECISION", "network":"LOCAL", "llm_capable": True},  # optional LLM phrasing
    "M": {"title":"REPORTS (XLSX + HTML)", "network":"LOCAL", "llm_capable": True},  # optional exec summary LLM
}

def _stage_range(start: str, end: str) -> List[str]:
    s = (start or "").upper().strip()
    e = (end or "").upper().strip()
    if s not in STAGE_ORDER:
        raise SystemExit(f"Unknown --start_stage {s}. Allowed: {', '.join(STAGE_ORDER)}")
    if e not in STAGE_ORDER:
        raise SystemExit(f"Unknown --end_stage {e}. Allowed: {', '.join(STAGE_ORDER)}")
    i0 = STAGE_ORDER.index(s)
    i1 = STAGE_ORDER.index(e)
    if i1 < i0:
        raise SystemExit("--end_stage must be >= --start_stage")
    return STAGE_ORDER[i0:i1+1]

def _print_stages() -> None:
    print("Stages A..M:")
    for st in STAGE_ORDER:
        inf = STAGE_INFO[st]
        net = inf["network"]
        llm = "optional" if inf["llm_capable"] else "no"
        print(f"  {st}: {inf['title']}  | net={net} | llm={llm}")

def _print_run_banner(args) -> None:
    print("\nWB Revival v2: Necromancer")
    print("Because humans insist on running pipelines at 23:00.\n")
    print(f"OUT: {args.out}")
    print(f"Stages: {args.start_stage}..{args.end_stage}  (resume={args.resume})")
    print(f"WB: dest={args.dest} locale={args.locale} timeout={args.wb_timeout}s sleep={args.sleep_s}s strict={args.strict}")
    llm_any = bool(args.use_llm_l or args.use_llm_m)
    print(f"LLM: {'ON' if llm_any else 'OFF'}  (L={args.use_llm_l}, M={args.use_llm_m})")
    if llm_any:
        print(f"  provider={args.llm_provider} model={args.llm_model} base={args.llm_api_base or '(default)'}")
    print()

def _print_stage_banner(stage: str, args) -> None:
    inf = STAGE_INFO.get(stage, {"title":"", "network":"", "llm_capable": False})
    net = inf["network"]
    llm_on = False
    if stage == "L" and args.use_llm_l:
        llm_on = True
    if stage == "M" and args.use_llm_m:
        llm_on = True

    print("=" * 80)
    print(f"[{stage}] {inf['title']}")
    print(f"Network: {net}")
    if net == "WB":
        print("VPN hint: OFF (WB stage)")
    else:
        if llm_on:
            print("VPN hint: ON (LLM stage)")
        else:
            print("VPN hint: doesn't matter (local-only)")
    if stage in ("L","M"):
        print(f"LLM: {'ON' if llm_on else 'OFF'}")
    print("=" * 80)

def _env_get(keys: List[str]) -> str:
    for k in keys:
        v = os.getenv(k, "").strip()
        if v:
            return v
    return ""

def _resolve_api_key(provider: str, api_key_arg: str) -> str:
    if api_key_arg and api_key_arg.strip():
        return api_key_arg.strip()
    prov = (provider or "").lower().strip()
    if prov == "openai":
        return _env_get(["OPENAI_API_KEY"])
    if prov == "openrouter":
        return _env_get(["OPENROUTER_API_KEY", "OPENAI_API_KEY"])
    return _env_get(["OPENROUTER_API_KEY", "OPENAI_API_KEY"])

def _resolve_input_path(inp: str) -> Path:
    if inp and inp.strip():
        return Path(inp).expanduser()
    # common defaults in current working dir
    for cand in ["WB_INPUT_64_FROM_POCKETS_POD.xlsx", "WB_старые_артикулы_AUDIT_TEMPLATE.xlsx"]:
        p = Path(cand)
        if p.exists() and p.is_file():
            return p.resolve()
    # fallback: first xlsx in cwd
    xs = sorted([p for p in Path.cwd().glob("*.xlsx") if p.is_file()])
    if xs:
        return xs[0].resolve()
    raise SystemExit("No --input provided and no .xlsx found in current directory.")

def _maybe_pause_before_llm(args, stages: List[str]) -> None:
    if not args.pause_before_llm:
        return
    # Only pause if LLM is actually enabled in any selected stage
    llm_stage_idx = None
    for i, st in enumerate(stages):
        if (st == "L" and args.use_llm_l) or (st == "M" and args.use_llm_m):
            llm_stage_idx = i
            break
    if llm_stage_idx is None:
        return
    if args.yes:
        print("[runner] pause_before_llm: auto-yes")
        return
    if not sys.stdin.isatty():
        print("[runner] pause_before_llm requested, but no TTY. Skipping pause.")
        return
    print("\n[runner] Next is an LLM stage. Turn VPN ON now if needed.")
    ans = input("Continue? [y/N]: ").strip().lower()
    if ans not in ("y", "yes"):
        raise SystemExit("Stopped by user.")

def build_cli_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="WB Revival v2 Necromancer runner")
    ap.add_argument("--list_stages", action="store_true", help="Print stages and exit")
    ap.add_argument("--out", required=True, help="Output directory (created if missing). Example: RUN_TEST")
    ap.add_argument("--input", default="", help="Input xlsx (Stage A). If empty: try default names, else first *.xlsx in cwd.")
    ap.add_argument("--sheet", default="INPUT_64", help="Sheet name (Stage A)")
    ap.add_argument("--start_stage", default="A", help="Start stage letter A..M")
    ap.add_argument("--end_stage", default="M", help="End stage letter A..M")
    ap.add_argument("--resume", action="store_true", help="Resume: skip nm_id already written in target JSONL")
    ap.add_argument("--expect_count", type=int, default=64, help="Expected SKU count (Stage A sanity). Use 0 to disable.")
    ap.add_argument("--dedupe", action="store_true", help="Dedupe nm_id in scope (Stage A)")
    ap.add_argument("--pause_before_llm", action="store_true", help="Pause once before first enabled LLM stage (VPN hint)")
    ap.add_argument("--yes", action="store_true", help="Auto-confirm pauses/prompts")

    # WB params
    ap.add_argument("--dest", type=int, default=-1257786, help="WB dest id")
    ap.add_argument("--locale", default="ru", help="WB locale/lang")
    ap.add_argument("--wb_timeout", type=int, default=30, help="WB HTTP timeout seconds")
    ap.add_argument("--sleep_s", type=float, default=0.4, help="Sleep between WB requests")
    ap.add_argument("--strict", action="store_true", help="Strict mode: raise on errors in WB stages")

    # LLM params (Stage L/M)
    ap.add_argument("--use_llm_l", action="store_true", help="Use LLM in Stage L (phrasing only, decision is rules-first)")
    ap.add_argument("--use_llm_m", action="store_true", help="Use LLM in Stage M (exec_summary only)")
    ap.add_argument("--llm_provider", default=os.getenv("LLM_PROVIDER", "openrouter"), help="openai|openrouter")
    ap.add_argument("--llm_model", default=os.getenv("LLM_MODEL", "openai/gpt-4o-mini"), help="Model id")
    ap.add_argument("--llm_api_base", default=os.getenv("LLM_API_BASE", ""), help="Override base URL (optional)")
    ap.add_argument("--llm_api_key", default="", help="API key (else from env)")
    ap.add_argument("--llm_timeout", type=int, default=60, help="LLM timeout seconds")
    ap.add_argument("--llm_max_tokens_l", type=int, default=1200, help="Stage L max tokens")
    ap.add_argument("--llm_max_tokens_m", type=int, default=1400, help="Stage M max tokens")
    ap.add_argument("--llm_temperature", type=float, default=0.2, help="LLM temperature")
    return ap

def run_selected_stages(args) -> None:
    if args.list_stages:
        _print_stages()
        return

    out_dir = Path(args.out).expanduser()
    ensure_dir(out_dir)

    stages = _stage_range(args.start_stage, args.end_stage)
    _print_run_banner(args)
    _maybe_pause_before_llm(args, stages)

    for st in stages:
        _print_stage_banner(st, args)

        if st == "A":
            input_xlsx = _resolve_input_path(args.input)
            stage_A(
                out_dir=out_dir,
                input_xlsx=input_xlsx,
                sheet=args.sheet,
                expect_count=(args.expect_count if args.expect_count and args.expect_count > 0 else None),
                dedupe=args.dedupe,
                resume=args.resume,
            )

        elif st == "B":
            stage_B_own_fetch(
                out_dir=out_dir,
                dest=args.dest,
                locale=args.locale,
                timeout=args.wb_timeout,
                sleep_s=args.sleep_s,
                deep_card=True,
                resume=args.resume,
                strict=args.strict,
            )

        elif st == "C":
            stage_C_intent_extract(out_dir=out_dir, resume=args.resume)

        elif st == "D":
            stage_D_query_build(out_dir=out_dir, resume=args.resume)

        elif st == "E":
            stage_E_serp_validate(out_dir=out_dir, timeout=args.wb_timeout, sleep_s=args.sleep_s, resume=args.resume)

        elif st == "F":
            stage_F_competitor_pool(out_dir=out_dir, resume=args.resume)

        elif st == "G":
            stage_G_competitor_lite_fetch(
                out_dir=out_dir,
                timeout=args.wb_timeout,
                sleep_s=args.sleep_s,
                resume=args.resume,
                strict=args.strict,
            )

        elif st == "H":
            stage_H_relevance_filter(out_dir=out_dir, resume=args.resume, strict=args.strict)

        elif st == "I":
            stage_I_market_pulse(
                out_dir=out_dir,
                timeout=args.wb_timeout,
                sleep_s=args.sleep_s,
                resume=args.resume,
                strict=args.strict,
            )

        elif st == "J":
            stage_J_supply_structure(out_dir=out_dir, resume=args.resume)

        elif st == "K":
            stage_K_cluster_verdicts(out_dir=out_dir, resume=args.resume)

        elif st == "L":
            api_key = _resolve_api_key(args.llm_provider, args.llm_api_key)
            stage_L_final_decision(
                out_dir=out_dir,
                resume=args.resume,
                use_llm_l=args.use_llm_l,
                llm_provider=args.llm_provider,
                llm_model=args.llm_model,
                api_key=api_key,
                llm_base_url=args.llm_api_base,
                llm_timeout=args.llm_timeout,
                llm_max_tokens=args.llm_max_tokens_l,
                llm_temperature=args.llm_temperature,
            )

        elif st == "M":
            api_key = _resolve_api_key(args.llm_provider, args.llm_api_key)
            stage_M_reports(
                out_dir=out_dir,
                resume=args.resume,
                use_llm_m=args.use_llm_m,
                llm_provider=args.llm_provider,
                llm_model=args.llm_model,
                api_key=api_key,
                llm_base_url=args.llm_api_base,
                llm_timeout=args.llm_timeout,
                llm_max_tokens=args.llm_max_tokens_m,
                llm_temperature=args.llm_temperature,
            )

        else:
            raise SystemExit(f"Stage not wired in runner: {st}")

    print("\nDONE. Output dir:", str(out_dir.resolve()))

def main() -> None:
    ap = build_cli_parser()
    args = ap.parse_args()
    run_selected_stages(args)

if __name__ == "__main__":
    main()
